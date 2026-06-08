"""Tests for round-tripping recommendations into the uploaded Amazon bulk sheet."""

from __future__ import annotations

import io
import unittest

import openpyxl

from sales_support_agent.services.advertising.bulk_sheets import build_bulk_workbook
from sales_support_agent.services.advertising.schema import Recommendation


_HEADER = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Keyword ID",
           "Campaign Name (Informational only)", "Ad Group Name (Informational only)",
           "Keyword Text", "Match Type", "Bid", "Impressions", "Clicks", "Spend", "Sales", "Orders", "Units"]


def _workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sponsored Products Campaigns"
    ws.append(_HEADER)

    def row(**k):
        ws.append([k.get(h, "") for h in _HEADER])

    row(Product="Sponsored Products", Entity="Ad Group",
        **{"Campaign ID": "C1", "Ad Group ID": "A1",
           "Campaign Name (Informational only)": "Brand", "Ad Group Name (Informational only)": "AG"})
    row(Product="Sponsored Products", Entity="Keyword",
        **{"Campaign ID": "C1", "Ad Group ID": "A1", "Keyword ID": "K1",
           "Campaign Name (Informational only)": "Brand", "Ad Group Name (Informational only)": "AG",
           "Keyword Text": "widget blue", "Match Type": "exact", "Bid": 1.20,
           "Impressions": 1000, "Clicks": 40, "Spend": 40, "Sales": 20, "Orders": 2, "Units": 2})
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rec(action, **bulk):
    bulk.setdefault("ad_type", "SP")
    bulk.setdefault("campaign_name", "Brand")
    bulk.setdefault("ad_group_name", "AG")
    bulk["action"] = action
    return Recommendation(category="x", title=f"{action}", is_bulk_actionable=True, bulk_row=bulk)


def _read(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Sponsored Products Campaigns"]
    hdr = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        out.append(dict(zip(hdr, r)))
    return out


class BulkRoundTripTest(unittest.TestCase):
    def test_set_bid_updates_row(self):
        recs = [_rec("set_bid", keyword_text="widget blue", match_type="exact", new_bid_cents=20)]
        res = build_bulk_workbook(_workbook(), recs)
        self.assertEqual(res.applied, 1)
        rows = _read(res.xlsx_bytes)
        kw = next(r for r in rows if r["Keyword Text"] == "widget blue")
        self.assertEqual(kw["Operation"], "Update")  # Amazon operation names are capitalized
        self.assertEqual(kw["Bid"], 0.20)

    def test_create_negative_appends_row(self):
        recs = [_rec("create_negative", keyword_text="cheap junk", match_type="negative exact")]
        res = build_bulk_workbook(_workbook(), recs)
        self.assertEqual(res.applied, 1)
        rows = _read(res.xlsx_bytes)
        neg = next(r for r in rows if r["Keyword Text"] == "cheap junk")
        self.assertEqual(neg["Entity"], "Negative Keyword")
        self.assertEqual(neg["Operation"], "create")
        self.assertEqual(neg["Match Type"], "negativeExact")
        self.assertEqual(neg["Campaign ID"], "C1")  # copied from sibling
        self.assertEqual(neg["Ad Group ID"], "A1")

    def test_create_keyword_appends_with_bid(self):
        recs = [_rec("create_keyword", keyword_text="widget green", match_type="exact", new_bid_cents=100)]
        res = build_bulk_workbook(_workbook(), recs)
        self.assertEqual(res.applied, 1)
        rows = _read(res.xlsx_bytes)
        kw = next(r for r in rows if r["Keyword Text"] == "widget green")
        self.assertEqual(kw["Entity"], "Keyword")
        self.assertEqual(kw["Operation"], "create")
        self.assertEqual(kw["Bid"], 1.00)

    def test_no_uploaded_file_skips(self):
        recs = [_rec("set_bid", keyword_text="widget blue", new_bid_cents=20)]
        res = build_bulk_workbook(b"", recs)
        self.assertEqual(res.applied, 0)
        self.assertEqual(res.skipped, 1)
        self.assertFalse(res.has_file)

    def test_unmatched_keyword_skipped(self):
        recs = [_rec("set_bid", keyword_text="does not exist", new_bid_cents=20)]
        res = build_bulk_workbook(_workbook(), recs)
        self.assertEqual(res.applied, 0)
        self.assertEqual(res.skipped, 1)

    def test_non_actionable_recs_ignored(self):
        rec = Recommendation(category="manual", title="manual", is_bulk_actionable=False)
        res = build_bulk_workbook(_workbook(), [rec])
        self.assertEqual(res.applied, 0)
        self.assertIsNone(res.xlsx_bytes)

    def test_untouched_rows_keep_blank_operation(self):
        recs = [_rec("create_negative", keyword_text="cheap junk", match_type="negative exact")]
        res = build_bulk_workbook(_workbook(), recs)
        rows = _read(res.xlsx_bytes)
        kw = next(r for r in rows if r["Keyword Text"] == "widget blue")
        self.assertIn(kw["Operation"], ("", None))  # not marked for an operation


if __name__ == "__main__":
    unittest.main()
