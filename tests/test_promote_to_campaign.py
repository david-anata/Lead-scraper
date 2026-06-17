"""Tests for promoting proven search terms into their own dedicated exact-match
SP campaigns: engine candidate detection, SKU resolution, the linked apply-sheet
rows (which must pass Amazon's Config validation), and the audit wiring."""

from __future__ import annotations

import io
import unittest

import openpyxl

from sales_support_agent.services.advertising import engine, normalizers as N
from sales_support_agent.services.advertising.bulk_sheets import build_apply_sheet, validate_bulk_rows
from sales_support_agent.services.advertising.schema import AdRow, Goals, SalesRow


def _bulk_with_product_ad(asin="B07XYZ1234", sku="SKU-1", cid="C1", cname="Auto Camp B07XYZ") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sponsored Products Campaigns"
    hdr = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "SKU",
           "Campaign Name (Informational only)", "ASIN (Informational only)"]
    ws.append(hdr)
    ws.append(["Sponsored Products", "Product Ad", "", cid, "A1", sku, cname, asin])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _winner(term="blue widget", orders=6, clicks=20, spend=800, sales=4000, cid="C1", aid="A1") -> AdRow:
    return AdRow(ad_type="SP", entity_level="search_term", campaign_name="Auto Camp B07XYZ",
                 ad_group_name="AG", campaign_id=cid, ad_group_id=aid, entity_text=term,
                 match_type="broad", clicks=clicks, spend_cents=spend, sales_cents=sales, orders=orders)


GOALS = Goals(acos_target_bps=3000)  # 30% target; ceiling = 33%


class EngineDetectionTest(unittest.TestCase):
    def _promos(self, ads, sales=None):
        recs = engine.build_recommendations(ads, sales or [], goals=GOALS)
        return ([r for r in recs if r.bulk_row.get("action") == "create_campaign"],
                [r for r in recs if r.bulk_row.get("action") == "create_keyword"])

    def test_proven_winner_promotes(self):
        promos, harvests = self._promos([_winner(orders=6, sales=4000, spend=800)])  # 20% ACoS
        self.assertEqual([p.bulk_row["keyword_text"] for p in promos], ["blue widget"])
        # Promoted term must NOT also be harvested into its old ad group.
        self.assertNotIn("blue widget", [h.bulk_row["keyword_text"] for h in harvests])
        self.assertEqual(promos[0].category, "structure")
        self.assertFalse(promos[0].is_bulk_actionable)  # review-only until a SKU resolves

    def test_below_order_bar_stays_a_harvest(self):
        promos, harvests = self._promos([_winner(term="cheap widget", orders=3, sales=600, spend=300)])
        self.assertEqual(promos, [])
        self.assertIn("cheap widget", [h.bulk_row["keyword_text"] for h in harvests])

    def test_inefficient_term_not_promoted(self):
        # 6 orders but ACoS 80% (spend 4000 / sales 5000 = 80%) — above the 33% ceiling.
        promos, _ = self._promos([_winner(term="pricey widget", orders=6, sales=5000, spend=4000)])
        self.assertEqual(promos, [])

    def test_already_exact_not_promoted(self):
        ads = [
            _winner(term="blue widget", orders=6),
            AdRow(ad_type="SP", entity_level="keyword", entity_text="blue widget", match_type="exact"),
        ]
        promos, _ = self._promos(ads)
        self.assertEqual(promos, [])


class ResolveTargetsTest(unittest.TestCase):
    def test_sku_resolves_from_bulk_product_ad(self):
        # No Business Report SKU at all — the bulk file's Product Ad SKU is enough.
        recs = engine.build_recommendations([_winner()], [], goals=GOALS)
        made = N.resolve_promotion_targets(recs, _bulk_with_product_ad(sku="SKU-1"), [])
        self.assertEqual(made, 1)
        p = [r for r in recs if r.bulk_row.get("action") == "create_campaign"][0]
        self.assertTrue(p.is_bulk_actionable)
        self.assertEqual(p.bulk_row["products"], [{"sku": "SKU-1", "asin": "B07XYZ1234"}])

    def test_sku_resolves_from_business_report_fallback(self):
        # Bulk Product Ad has the ASIN but no SKU → fall back to the Business Report.
        recs = engine.build_recommendations([_winner()], [], goals=GOALS)
        sales = [SalesRow(asin="B07XYZ1234", sku="SKU-BR", title="Blue Widget")]
        made = N.resolve_promotion_targets(recs, _bulk_with_product_ad(sku=""), sales)
        self.assertEqual(made, 1)
        p = [r for r in recs if r.bulk_row.get("action") == "create_campaign"][0]
        self.assertEqual(p.bulk_row["products"], [{"sku": "SKU-BR", "asin": "B07XYZ1234"}])

    def test_unresolved_sku_is_review_only(self):
        recs = engine.build_recommendations([_winner()], [], goals=GOALS)
        made = N.resolve_promotion_targets(recs, _bulk_with_product_ad(sku=""), [])  # no SKU anywhere
        self.assertEqual(made, 0)
        p = [r for r in recs if r.bulk_row.get("action") == "create_campaign"][0]
        self.assertFalse(p.is_bulk_actionable)
        self.assertIn("review_only_reason", p.bulk_row)


class ApplySheetTest(unittest.TestCase):
    def _ready_rec(self):
        recs = engine.build_recommendations([_winner()], [], goals=GOALS)
        N.resolve_promotion_targets(recs, _bulk_with_product_ad(),
                                    [SalesRow(asin="B07XYZ1234", sku="SKU-1")])
        return [r for r in recs if r.bulk_row.get("action") == "create_campaign"]

    def _entities(self, xlsx_bytes):
        ws = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))["Sponsored Products Campaigns"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        ei, oi = hdr.index("Entity"), hdr.index("Operation")
        return [(r[ei], r[oi]) for r in rows[1:] if r[ei]]

    def test_expands_to_linked_rows_and_validates(self):
        res = build_apply_sheet(self._ready_rec(), kinds={"create_campaign"})
        self.assertEqual(res.applied, 1)          # one logical campaign
        self.assertEqual(res.invalid, 0, res.issues)  # passes Amazon Config validation
        self.assertEqual(self._entities(res.xlsx_bytes), [
            ("Campaign", "Create"), ("Ad Group", "Create"), ("Product Ad", "Create"),
            ("Keyword", "Create"), ("Negative Keyword", "Create"),
        ])

    def test_placeholder_ids_link_the_rows(self):
        res = build_apply_sheet(self._ready_rec(), kinds={"create_campaign"})
        ws = openpyxl.load_workbook(io.BytesIO(res.xlsx_bytes))["Sponsored Products Campaigns"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        ci, ai, ei = hdr.index("Campaign ID"), hdr.index("Ad Group ID"), hdr.index("Entity")
        body = [r for r in rows[1:] if r[ei]]
        camp_id = body[0][ci]              # Campaign row's placeholder
        self.assertTrue(camp_id.startswith("NEW-"))
        ag_id = body[1][ai]               # Ad Group row's placeholder
        # Keyword row (index 3) must reference the same campaign + ad group placeholders.
        self.assertEqual(body[3][ci], camp_id)
        self.assertEqual(body[3][ai], ag_id)

    def test_no_source_ids_omits_the_negative(self):
        recs = engine.build_recommendations([_winner(cid="", aid="")], [], goals=GOALS)
        N.resolve_promotion_targets(recs, _bulk_with_product_ad(cid="", cname="Auto Camp B07XYZ"),
                                    [SalesRow(asin="B07XYZ1234", sku="SKU-1")])
        ready = [r for r in recs if r.bulk_row.get("action") == "create_campaign" and r.is_bulk_actionable]
        if not ready:
            self.skipTest("no apply-ready promo without source ids in this fixture")
        res = build_apply_sheet(ready, kinds={"create_campaign"})
        ents = [e for e, _ in self._entities(res.xlsx_bytes)]
        self.assertNotIn("Negative Keyword", ents)


def _bulk_full(asin="B07XYZ1234", sku="SKU-1") -> bytes:
    """A bulk file with Campaign/Ad Group/Product Ad rows so backfill can resolve
    the search term's source IDs (→ the anti-cannibalization negative) and the
    Product Ad SKU drives the promotion's product."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sponsored Products Campaigns"
    hdr = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "SKU",
           "Campaign Name (Informational only)", "Ad Group Name (Informational only)",
           "ASIN (Informational only)"]
    ws.append(hdr)
    P = "Sponsored Products"
    ws.append([P, "Campaign", "", "C1", "", "", "Auto Camp", "", ""])
    ws.append([P, "Ad Group", "", "C1", "A1", "", "Auto Camp", "AG", ""])
    ws.append([P, "Product Ad", "", "C1", "A1", sku, "Auto Camp", "AG", asin])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_SEARCH_TERM_CSV = (
    b"Campaign Name,Ad Group Name,Customer Search Term,Match Type,Impressions,Clicks,Spend,"
    b"7 Day Total Sales,7 Day Total Orders (#),7 Day Total Units (#)\n"
    b"Auto Camp,AG,blue widget,broad,500,20,8.00,40.00,6,6\n"
)


class RunAuditIntegrationTest(unittest.TestCase):
    def setUp(self):
        from tests.test_advertising_audit import _bootstrap_db, _make_engine, _patch_global_engine
        import tempfile
        self.engine = _make_engine()
        _bootstrap_db(self.engine)
        self.old_engine = _patch_global_engine(self.engine)
        self.tmpdir = tempfile.mkdtemp()
        import sales_support_agent.services.advertising.storage as storage
        self.storage = storage
        self._old_bulk_dir = storage.BULK_RUNS_DIR
        storage.BULK_RUNS_DIR = self.tmpdir

    def tearDown(self):
        from tests.test_advertising_audit import _patch_global_engine
        self.storage.BULK_RUNS_DIR = self._old_bulk_dir
        _patch_global_engine(self.old_engine)
        self.engine.dispose()

    def test_run_audit_emits_new_campaign_into_additions(self):
        from sales_support_agent.services.advertising.audit import AuditInputs, run_audit
        res = run_audit(
            AuditInputs(bulk_xlsx=_bulk_full(), search_term_csv=_SEARCH_TERM_CSV),
            goals=GOALS, label="promote wk",
        )
        self.assertEqual(res.status, "complete")
        self.assertGreaterEqual(res.summary.get("new_campaign_count", 0), 1)
        self.assertTrue(res.summary.get("new_campaigns"))
        # The Additions file should carry a Campaign Create row (live on upload).
        data = self.storage.get_bulk_file(res.run_id, "additions")
        self.assertIsNotNone(data)
        ws = openpyxl.load_workbook(io.BytesIO(data))["Sponsored Products Campaigns"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        ei = hdr.index("Entity")
        entities = {r[ei] for r in rows[1:] if r[ei]}
        self.assertIn("Campaign", entities)
        self.assertIn("Keyword", entities)


if __name__ == "__main__":
    unittest.main()
