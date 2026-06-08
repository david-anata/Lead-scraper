"""Tests for brand focus + the growth-plan workbook deliverable."""

from __future__ import annotations

import io
import unittest

import openpyxl

from sales_support_agent.services.advertising.brand import detect_brand_candidates, filter_by_brand
from sales_support_agent.services.advertising.deliverable import build_growth_plan
from sales_support_agent.services.advertising.schema import AdRow, Goals, SalesRow


def _ad(level, *, campaign="", asin="", text="", spend=0, sales=0, orders=0, clicks=0, ad_type="SP"):
    raw = {"Advertised product ID": asin} if asin else {}
    return AdRow(ad_type=ad_type, entity_level=level, campaign_name=campaign, ad_group_name="",
                entity_text=text or asin, impressions=clicks * 20, clicks=clicks,
                spend_cents=spend, sales_cents=sales, orders=orders, units=orders, raw=raw)


class BrandFilterTest(unittest.TestCase):
    def setUp(self):
        # Zantrex campaign named by ASIN (no brand word), plus a Serovital one.
        self.ads = [
            _ad("product_ad", campaign="P_..._B07NXN4F7X_ATM", asin="B07NXN4F7X", spend=20000, sales=31000, orders=10),
            _ad("search_term", campaign="P_..._B07NXN4F7X_ATM", text="skinny stix", spend=1500, sales=6000, orders=4, clicks=15),
            _ad("product_ad", campaign="Quartile Serovital - NB", asin="B0FN5SCC1T", spend=9000, sales=20000, orders=5),
        ]
        self.sales = [
            SalesRow(asin="B07NXN4F7X", title="Zantrex SkinnyStix Energy Powder", sessions=60000, units=3117,
                     ordered_product_sales_cents=11144718, conversion_bps=519),
            SalesRow(asin="B0FN5SCC1T", title="Serovital SmileCare Toothpaste", sessions=900, units=20,
                     ordered_product_sales_cents=200000),
        ]

    def test_asin_aware_filter_keeps_asin_named_campaigns(self):
        ads, sales = filter_by_brand(self.ads, self.sales, "Zantrex")
        # The product_ad (ASIN match) AND the search term in that campaign are kept,
        # even though neither name contains "Zantrex".
        self.assertEqual(len(ads), 2)
        self.assertEqual(len(sales), 1)
        self.assertEqual(sum(r.spend_cents for r in ads if r.entity_level == "product_ad"), 20000)
        # Serovital is excluded.
        self.assertTrue(all("Serovital" not in r.campaign_name for r in ads))

    def test_mixed_cross_brand_campaign_excluded(self):
        from sales_support_agent.services.advertising.brand import mixed_campaigns
        ads = [
            _ad("product_ad", campaign="Zantrex_clean", asin="B07NXN4F7X", spend=1000, sales=2000, orders=2),
            # a catch-all campaign advertising Zantrex AND another brand together
            _ad("product_ad", campaign="Portfolio_mixed", asin="B07NXN4F7X", spend=500, sales=900, orders=1),
            _ad("product_ad", campaign="Portfolio_mixed", asin="B0SEROVITAL", spend=400, sales=800, orders=1),
            _ad("search_term", campaign="Portfolio_mixed", text="junk", spend=900, sales=0, orders=0, clicks=20),
        ]
        sales = [
            SalesRow(asin="B07NXN4F7X", title="Zantrex SkinnyStix", units=10, ordered_product_sales_cents=20000),
            SalesRow(asin="B0SEROVITAL", title="Serovital HGH", units=5, ordered_product_sales_cents=10000),
        ]
        mixed = mixed_campaigns(ads, sales, "Zantrex")
        self.assertIn("Portfolio_mixed", mixed)
        self.assertNotIn("Zantrex_clean", mixed)
        kept, _ = filter_by_brand(ads, sales, "Zantrex")
        mixed_kept = [r for r in kept if r.campaign_name == "Portfolio_mixed"]
        # from the mixed campaign, only the brand's product-ad row survives (so its
        # spend still counts) — the targeting row that could spawn an edit is gone.
        self.assertTrue(mixed_kept and all(r.entity_level == "product_ad" for r in mixed_kept))
        self.assertFalse(any(r.entity_text == "junk" for r in kept))  # edit-risk search term dropped
        self.assertTrue(any(r.campaign_name == "Zantrex_clean" for r in kept))

    def test_single_brand_account_excludes_nothing(self):
        from sales_support_agent.services.advertising.brand import mixed_campaigns
        ads = [_ad("product_ad", campaign="C", asin="B07NXN4F7X", spend=1000, sales=2000, orders=2)]
        sales = [SalesRow(asin="B07NXN4F7X", title="Zantrex SkinnyStix", units=10, ordered_product_sales_cents=20000)]
        self.assertEqual(mixed_campaigns(ads, sales, "Zantrex"), set())  # no other-brand ASINs -> no-op

    def test_blank_brand_returns_all(self):
        ads, sales = filter_by_brand(self.ads, self.sales, "")
        self.assertEqual(len(ads), 3)
        self.assertEqual(len(sales), 2)

    def test_detect_primary_brand_multiword(self):
        from sales_support_agent.services.advertising.brand import detect_primary_brand
        sales = [SalesRow(asin=f"B{i}", title=t) for i, t in enumerate([
            "Number 4 Hydrating Shampoo, Color Safe",
            "Number 4 Super Comb Leave-In Conditioner",
            "Number 4 Hair Oil, Anti Frizz Serum",
            "Number 4 Smoothing Hair Balm",
        ])]
        self.assertEqual(detect_primary_brand(sales), "Number 4")

    def test_detect_primary_brand_blank_for_multibrand(self):
        from sales_support_agent.services.advertising.brand import detect_primary_brand
        sales = [SalesRow(asin="A", title="Zantrex SkinnyStix"),
                 SalesRow(asin="B", title="Serovital HGH"),
                 SalesRow(asin="C", title="Nugenix Total T"),
                 SalesRow(asin="D", title="Zantrex Black")]
        self.assertEqual(detect_primary_brand(sales), "")  # no dominant brand → 'Full account'

    def test_detect_candidates_skips_codes_and_asins(self):
        ads = [
            _ad("campaign", campaign="Quartile Zantrex - Non Branded - Skinnystix"),
            _ad("campaign", campaign="Quartile Zantrex - Branded"),
            _ad("campaign", campaign="P_APRQ8M711018_B07NXN4F7X_ATM"),
        ]
        cands = detect_brand_candidates(ads)
        self.assertIn("Zantrex", cands)
        self.assertNotIn("APRQ8M711018", cands)
        self.assertNotIn("B07NXN4F7X", cands)


class GrowthPlanTest(unittest.TestCase):
    def setUp(self):
        self.ads = [
            _ad("product_ad", campaign="P_B07NXN4F7X", asin="B07NXN4F7X", spend=2000000, sales=3100000, orders=100),
            _ad("search_term", campaign="P_B07NXN4F7X", text="cheap junk", spend=5000, sales=0, orders=0, clicks=25),
            _ad("search_term", campaign="P_B07NXN4F7X", text="skinny stix", spend=1500, sales=6000, orders=4, clicks=15),
        ]
        self.sales = [
            SalesRow(asin="B07NXN4F7X", title="Zantrex SkinnyStix", sessions=60000, units=3117,
                     ordered_product_sales_cents=11144718, conversion_bps=519, buy_box_pct_bps=9513),
        ]
        self.goals = Goals(revenue_target_cents=45000000, acos_target_bps=3000, tacos_target_bps=1800)
        from sales_support_agent.services.advertising.engine import build_recommendations
        self.recs = build_recommendations(self.ads, self.sales, [], [], self.goals)

    def test_workbook_has_all_tabs(self):
        data = build_growth_plan(brand="Zantrex", summary={"total_sales_cents": 11144718, "ad_spend_cents": 2006500,
                                 "tacos_bps": 1800, "acos_bps": 6400, "total_units": 3117, "external_spend_cents": 0,
                                 "blended_tacos_bps": 1800},
                                 recommendations=self.recs, ad_rows=self.ads, sales_rows=self.sales,
                                 goals=self.goals, narrative="Test read.")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        self.assertEqual(wb.sheetnames,
                         ["Exec Brief", "Burn List", "ASIN Scorecard", "Campaign Actions",
                          "Negatives to Add", "Revenue Bridge", "Data Requests"])

    def test_asin_scorecard_joins_org_and_ad(self):
        data = build_growth_plan(brand="Zantrex", summary={"total_sales_cents": 11144718, "ad_spend_cents": 2006500,
                                 "tacos_bps": 1800}, recommendations=self.recs, ad_rows=self.ads,
                                 sales_rows=self.sales, goals=self.goals)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["ASIN Scorecard"]
        rows = list(ws.iter_rows(values_only=True))
        header = next(r for r in rows if r and r[0] == "ASIN")
        data_row = rows[rows.index(header) + 1]
        self.assertIn("ASIN", header)
        self.assertEqual(data_row[0], "B07NXN4F7X")
        # org sales (dollars) and ad spend both present on one row
        self.assertAlmostEqual(data_row[2], 111447.18, places=2)
        self.assertAlmostEqual(data_row[7], 20000.0, places=2)  # ad spend $20,000

    def test_strategic_read_filled_without_api_key(self):
        from sales_support_agent.services.advertising.llm import build_deterministic_read
        text = build_deterministic_read(
            {"total_sales_cents": 29682674, "blended_tacos_bps": 1520, "acos_bps": 5370,
             "gap": {"revenue_gap_cents": 15317326, "revenue_target_cents": 45000000, "revenue_attainment_bps": 6600}},
            self.recs, self.goals)
        self.assertIn("Revenue", text)
        self.assertNotIn("API_KEY", text)  # never a 'set the key' placeholder
        self.assertNotIn("unavailable", text)

    def test_conditional_formatting_applied(self):
        data = build_growth_plan(brand="Z", summary={"total_sales_cents": 11144718, "ad_spend_cents": 2006500,
                                 "tacos_bps": 1800, "acos_bps": 6400, "gap": {}}, recommendations=self.recs,
                                 ad_rows=self.ads, sales_rows=self.sales, goals=self.goals)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        rules = sum(len(r.rules) for r in wb["ASIN Scorecard"].conditional_formatting)
        self.assertGreater(rules, 0)  # scorecard has color scales + verdict highlights

    def test_strategic_read_surfaces_scope_and_window_warning(self):
        from sales_support_agent.services.advertising.llm import build_deterministic_read
        text = build_deterministic_read(
            {"brand": "Zantrex", "brand_asin_count": 15, "excluded_mixed_campaigns": 3,
             "data_windows": ["May 07 - May 28", "May 16 - Jun 03"], "total_sales_cents": 1},
            self.recs, self.goals)
        self.assertIn("Scoped to 15 Zantrex", text)
        self.assertIn("3 cross-brand campaign", text)
        self.assertIn("different date windows", text)  # the data-sanity warning fires

    def test_no_window_warning_when_consistent(self):
        from sales_support_agent.services.advertising.llm import build_deterministic_read
        text = build_deterministic_read(
            {"brand": "Z", "brand_asin_count": 5, "data_windows": ["May 07 - May 28"], "total_sales_cents": 1},
            self.recs, self.goals)
        self.assertNotIn("different date windows", text)

    def test_data_requests_has_where_column(self):
        data = build_growth_plan(brand="Z", summary={"total_sales_cents": 1}, recommendations=self.recs,
                                 ad_rows=self.ads, sales_rows=self.sales, goals=self.goals)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        rows = list(wb["Data Requests"].iter_rows(values_only=True))
        header = next(r for r in rows if r and r[0] == "#")
        self.assertIn("Where / how to get it", header)
        # the row after the header carries real directions (Amazon / Seller Central path)
        first = rows[rows.index(header) + 1]
        self.assertTrue(any(tok in str(first[2]) for tok in ("Amazon Ads Console", "Seller Central", "ASIN")))

    def test_data_requests_flags_cogs_when_absent(self):
        data = build_growth_plan(brand="Zantrex", summary={"total_sales_cents": 1}, recommendations=self.recs,
                                 ad_rows=self.ads, sales_rows=self.sales, goals=self.goals, has_cogs=False)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        text = " ".join(str(v) for row in wb["Data Requests"].iter_rows(values_only=True) for v in row if v)
        self.assertIn("COGS", text)


if __name__ == "__main__":
    unittest.main()
