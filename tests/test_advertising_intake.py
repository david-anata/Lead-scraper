"""Tests for the auto-detecting mass-upload intake (sniff + route)."""

from __future__ import annotations

import io
import unittest

import openpyxl

from sales_support_agent.services.advertising import intake as I


def _bulk_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sponsored Products Campaigns"
    ws.append(["Product", "Entity", "Operation", "Keyword Text", "Bid", "Impressions", "Clicks", "Spend"])
    ws.append(["Sponsored Products", "Keyword", "", "widget", 1.0, 100, 5, 5.0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_SEARCH_TERM = b"Campaign Name,Customer Search Term,Impressions,Clicks,Spend,7 Day Total Orders (#)\nC,blue,100,5,5,0\n"
_BUSINESS = b"(Child) ASIN,Sessions - Total,Units Ordered,Ordered Product Sales\nB001,900,22,$320\n"
_SQP = b"Search Query,Search Query Volume,Impressions: Total Count,Purchases: ASIN Share %\nblue,1000,5000,20%\n"
_DSP = b"Campaign Name,Impressions,Clicks,Total Cost,Total Sales\nBrand,9000,10,$300,$1500\n"
_EXTERNAL = b"Channel,Amount,Note\nMeta,$1000,prospecting\n"


class SniffTest(unittest.TestCase):
    def test_bulk_xlsx(self):
        self.assertEqual(I.sniff_kind("MyBulk.xlsx", _bulk_xlsx()), I.KIND_BULK)

    def test_search_term(self):
        self.assertEqual(I.sniff_kind("st.csv", _SEARCH_TERM), I.KIND_ADS_REPORT)

    def test_new_console_ads_report(self):
        # New Amazon reporting console format (Total cost / Purchases / Sales).
        csv = (
            b"Campaign name,Ad group name,Search term,Impressions,Clicks,CTR,Total cost,Purchases,Sales,Units sold\n"
            b"Camp,AG,protein packets,13,2,15%,1.21,0,0.00,0\n"
        )
        self.assertEqual(I.sniff_kind("Search_term_06_04.csv", csv), I.KIND_ADS_REPORT)

    def test_portfolio_xlsx_not_bulk(self):
        # A non-bulk workbook (no Sponsored ... Campaigns sheet) must not be "bulk".
        wb = openpyxl.Workbook()
        wb.active.title = "Portfolio Trends"
        wb.active.append(["Portfolio", "Spend"])
        buf = io.BytesIO()
        wb.save(buf)
        self.assertEqual(I.sniff_kind("Zantrex Portfolio Trends.xlsx", buf.getvalue()), I.KIND_UNKNOWN)

    def test_dsp_by_filename(self):
        self.assertEqual(I.sniff_kind("DSP_report.csv", _DSP), I.KIND_DSP)

    def test_business_report(self):
        self.assertEqual(I.sniff_kind("BusinessReport.csv", _BUSINESS), I.KIND_BUSINESS)

    def test_sqp(self):
        self.assertEqual(I.sniff_kind("sqp.csv", _SQP), I.KIND_SQP)

    def test_dsp(self):
        self.assertEqual(I.sniff_kind("dsp.csv", _DSP), I.KIND_DSP)

    def test_external(self):
        self.assertEqual(I.sniff_kind("spend.csv", _EXTERNAL), I.KIND_EXTERNAL)

    def test_unknown(self):
        self.assertEqual(I.sniff_kind("notes.csv", b"foo,bar\n1,2\n"), I.KIND_UNKNOWN)

    def test_empty(self):
        self.assertEqual(I.sniff_kind("x.csv", b""), I.KIND_UNKNOWN)

    def test_legacy_xlsx_ads_report_is_classified_and_parsed(self):
        # Amazon "data export" Ads reports (Spend / 7-Day Total Sales / Advertised
        # ASIN) are commonly .xlsx — they must NOT be dropped as unknown workbooks
        # (that bug zeroed ad sales + suppressed the apply sheet).
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sponsored_Products_Advertised_p"
        ws.append(["Start Date", "Campaign Name", "Ad Group Name", "Advertised SKU",
                   "Advertised ASIN", "Impressions", "Clicks", "Spend",
                   "7 Day Total Sales ", "7 Day Total Orders (#)", "7 Day Total Units (#)"])
        ws.append(["2026-05-01", "Camp A", "AG 1", "SKU-1", "B005GEZGSQ",
                   1000, 40, 25.50, 120.00, 6, 6])
        buf = io.BytesIO(); wb.save(buf); data = buf.getvalue()

        self.assertEqual(I.sniff_kind("Sponsored_Products_Advertised_product_report.xlsx", data),
                         I.KIND_ADS_REPORT)
        from sales_support_agent.services.advertising.normalizers import normalize_ads_report_csv
        rows = normalize_ads_report_csv(data)
        pa = [r for r in rows if r.entity_level == "product_ad"]
        self.assertTrue(pa)
        self.assertEqual(pa[0].sales_cents, 12000)   # "7 Day Total Sales " (trailing space)
        self.assertEqual(pa[0].spend_cents, 2550)    # "Spend" alias
        self.assertEqual(pa[0].raw.get("Advertised product ID"), "B005GEZGSQ")  # ASIN canonicalized

    def test_sb_campaign_report_not_mistaken_for_bulk(self):
        # A single-sheet SB *Campaign report* has a sheet name with
        # "sponsored"+"campaign" but is a performance export, not a bulk file.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sponsored_Brands_Campaign_repor"
        ws.append(["Start Date", "Campaign Name", "Impressions", "Clicks", "Spend",
                   "14 Day Total Sales ", "14 Day Total Orders (#)"])
        ws.append(["2026-05-01", "SB Camp", 5000, 90, 80.00, 400.00, 12])
        buf = io.BytesIO(); wb.save(buf); data = buf.getvalue()
        self.assertFalse(I._xlsx_is_bulk(data))
        self.assertEqual(I.sniff_kind("Sponsored_Brands_Campaign_report.xlsx", data),
                         I.KIND_ADS_REPORT)

    def test_real_bulk_still_detected_by_entity_operation_header(self):
        self.assertTrue(I._xlsx_is_bulk(_bulk_xlsx()))


class RouteTest(unittest.TestCase):
    def test_routes_mixed_batch(self):
        inputs, report = I.route_files([
            ("bulk.xlsx", _bulk_xlsx()),
            ("st.csv", _SEARCH_TERM),
            ("biz.csv", _BUSINESS),
            ("random.csv", b"a,b\n1,2\n"),
        ])
        self.assertIsNotNone(inputs.bulk_xlsx)
        self.assertEqual(len(inputs.ads_report_csvs), 1)
        self.assertIsNotNone(inputs.business_report_csv)
        self.assertIn(I.KIND_BULK, report.detected)
        self.assertEqual(report.ignored, ["random.csv"])
        self.assertEqual(report.missing_core(), [])  # ads report + business present

    def test_missing_core_reported(self):
        # Only an ads report -> Business Report still missing (bulk is NOT core).
        _, report = I.route_files([("st.csv", _SEARCH_TERM)])
        missing = report.missing_core()
        self.assertIn(I.KIND_BUSINESS, missing)
        self.assertNotIn(I.KIND_ADS_REPORT, missing)
        self.assertNotIn(I.KIND_BULK, missing)

    def test_multiple_ads_reports_kept_separate(self):
        extra = b"Campaign name,Ad group name,Advertised product SKU,Impressions,Clicks,Total cost,Purchases,Sales,Units sold\nC,AG,SKU1,80,4,2.00,1,40.00,1\n"
        inputs, report = I.route_files([("search.csv", _SEARCH_TERM), ("product.csv", extra)])
        # Each ad report is parsed independently (not merged into one).
        self.assertEqual(len(inputs.ads_report_csvs), 2)
        from sales_support_agent.services.advertising.normalizers import normalize_ads_report_csv
        levels = set()
        for c in inputs.ads_report_csvs:
            levels |= {r.entity_level for r in normalize_ads_report_csv(c)}
        self.assertIn("search_term", levels)
        self.assertIn("product_ad", levels)

    def test_second_bulk_file_ignored(self):
        inputs, report = I.route_files([("a.xlsx", _bulk_xlsx()), ("b.xlsx", _bulk_xlsx())])
        self.assertIsNotNone(inputs.bulk_xlsx)
        self.assertIn("b.xlsx", report.ignored)

    def test_summary_human_readable(self):
        _, report = I.route_files([("bulk.xlsx", _bulk_xlsx()), ("st.csv", _SEARCH_TERM)])
        s = report.summary()
        self.assertIn("Detected", s)
        self.assertIn("Missing", s)  # business report still missing


if __name__ == "__main__":
    unittest.main()
