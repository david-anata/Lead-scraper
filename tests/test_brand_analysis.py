"""Deterministic tests for the Brand Analysis parsing + scoring pipeline.

The grade must be reproducible given identical inputs, so these never touch the
LLM (use_llm=False) or the database — they exercise intake -> scoring ->
confidence and the .docx export directly.
"""

from __future__ import annotations

import io
import os
import unittest

try:
    from sales_support_agent.services.brand_analysis import confidence as confidence_mod
    from sales_support_agent.services.brand_analysis import intake as intake_mod
    from sales_support_agent.services.brand_analysis import scoring as scoring_mod
    from sales_support_agent.services.brand_analysis.report import build_report
    from sales_support_agent.services.brand_analysis.schema import (
        Metrics,
        PeriodFinancials,
        benchmarks_for,
    )
    DEPS_AVAILABLE = True
except ModuleNotFoundError as exc:  # sqlalchemy / openpyxl missing in a thin env
    if exc.name not in {"sqlalchemy", "openpyxl"}:
        raise
    DEPS_AVAILABLE = False

_FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures", "brand_analysis")


def _load(name: str) -> tuple[str, bytes]:
    with open(os.path.join(_FIXTURES, name), "rb") as fh:
        return (name, fh.read())


@unittest.skipUnless(DEPS_AVAILABLE, "brand_analysis deps required")
class IntakeMetricsTests(unittest.TestCase):
    def setUp(self) -> None:
        self.report = build_report(
            [_load("acme_pnl_2024.csv"), _load("acme_balance_sheet.csv")],
            brand="Acme",
            category="dtc",
            use_llm=False,
        )

    def test_periods_detected(self) -> None:
        self.assertTrue(self.report.has_yoy)
        self.assertEqual(self.report.period_current_label, "FY 2024")
        self.assertEqual(self.report.period_prior_label, "FY 2023")

    def test_derived_metrics_exact(self) -> None:
        m = self.report.current
        self.assertEqual(m.net_revenue_cents, 96_000_000)        # $960,000
        self.assertEqual(m.cogs_cents, 37_000_000)
        self.assertEqual(m.product_gross_profit_cents, 59_000_000)
        self.assertEqual(m.product_gm_bps, 6146)                 # 61.46%
        self.assertEqual(m.marketing_pct_bps, 2604)              # 26.04%
        self.assertAlmostEqual(m.blended_mer, 3.84, places=2)
        self.assertEqual(m.contribution_margin_bps, 6146)        # reported GP / net rev
        self.assertEqual(m.net_margin_bps, 990)                  # 9.90%
        self.assertEqual(m.discount_rate_bps, 1500)              # 15%
        self.assertEqual(m.return_rate_bps, 500)                 # 5%
        self.assertEqual(self.report.yoy_revenue_growth_bps, 1852)  # +18.52%

    def test_media_and_balance_sections(self) -> None:
        self.assertEqual(self.report.media_mix.get("meta"), 15_000_000)
        self.assertEqual(self.report.media_mix.get("google"), 7_000_000)
        self.assertTrue(self.report.related_party_flag)          # intercompany line
        labels = {l for l, _ in self.report.balance_sheet}
        self.assertIn("Total assets", labels)
        self.assertIn("Intercompany balances", labels)

    def test_scorecard_is_deterministic(self) -> None:
        again = build_report(
            [_load("acme_pnl_2024.csv"), _load("acme_balance_sheet.csv")],
            brand="Acme", category="dtc", use_llm=False,
        )
        self.assertEqual(self.report.scorecard.score_100, again.scorecard.score_100)
        self.assertEqual(self.report.scorecard.letter, again.scorecard.letter)
        # Pin the headline grade for this fixture (financial track only — brand
        # moved to the separate Brand & Social track, so 7 weighted dimensions).
        self.assertEqual(self.report.scorecard.letter, "C")
        self.assertEqual(self.report.scorecard.score_100, 77)
        self.assertEqual(len(self.report.scorecard.dimensions), 7)

    def test_dimension_grades(self) -> None:
        by_key = {d.key: d for d in self.report.scorecard.dimensions}
        self.assertEqual(by_key["revenue"].letter, "B")          # +18.5% YoY
        self.assertEqual(by_key["balance"].letter, "D")          # intercompany + related party
        self.assertEqual(by_key["contribution"].letter, "A")
        weights = round(sum(d.weight for d in self.report.scorecard.dimensions), 4)
        self.assertEqual(weights, 1.0)

    def test_red_flags_surface(self) -> None:
        titles = [f.title for f in self.report.red_flags]
        self.assertIn("Related-party items present", titles)


@unittest.skipUnless(DEPS_AVAILABLE, "brand_analysis deps required")
class ConfidenceTests(unittest.TestCase):
    def test_missing_data_when_inputs_absent(self) -> None:
        report = build_report([_load("acme_pnl_2024.csv"), _load("acme_balance_sheet.csv")],
                              brand="Acme", use_llm=False)
        self.assertFalse(report.data_sufficient)
        self.assertEqual(report.confidence, "Medium")
        self.assertTrue(report.missing_data)

    def test_data_sufficient_path(self) -> None:
        full = PeriodFinancials(
            gross_sales_cents=120_000_000, discounts_cents=18_000_000, returns_cents=6_000_000,
            net_revenue_cents=96_000_000, cogs_cents=37_000_000, marketing_total_cents=25_000_000,
            reported_gross_profit_cents=59_000_000, net_earnings_cents=9_500_000, opex_cents=18_000_000,
            total_assets_cents=80_000_000, total_equity_cents=43_000_000,
            owned_channel_revenue_cents=20_000_000,
            new_customer_revenue_cents=50_000_000, returning_customer_revenue_cents=46_000_000,
            marketing_by_channel={"meta": 15_000_000, "google": 10_000_000},
        )
        result = confidence_mod.evaluate(full, has_yoy=True)
        self.assertTrue(result["data_sufficient"])
        self.assertEqual(result["confidence"], "High")
        self.assertEqual(result["missing_short"], [])


@unittest.skipUnless(DEPS_AVAILABLE, "brand_analysis deps required")
class SinglePeriodTests(unittest.TestCase):
    def test_single_period_mode(self) -> None:
        csv = b"Line Item,Amount\nNet Revenue,500000\nCOGS,200000\nTotal Marketing,120000\nNet Earnings,40000\n"
        report = build_report([("brand_pnl.csv", csv)], brand="Solo", use_llm=False)
        self.assertFalse(report.has_yoy)
        self.assertIsNone(report.yoy_revenue_growth_bps)
        revenue_dim = next(d for d in report.scorecard.dimensions if d.key == "revenue")
        self.assertIn("No prior-year", revenue_dim.reason)


@unittest.skipUnless(DEPS_AVAILABLE, "brand_analysis deps required")
class TrajectoryTests(unittest.TestCase):
    def test_declining_margin_downgrades(self) -> None:
        bm = benchmarks_for("dtc")
        # Same current value; only the prior differs (stable vs much higher).
        stable_letter, _ = scoring_mod._grade_margin_like(1000, *bm.net_margin_bps, 1000, "Net margin")
        declining_letter, reason = scoring_mod._grade_margin_like(1000, *bm.net_margin_bps, 1400, "Net margin")
        self.assertIn("Declining", reason)
        order = ["A", "B", "C", "D", "F"]
        self.assertGreater(order.index(declining_letter), order.index(stable_letter))


@unittest.skipUnless(DEPS_AVAILABLE, "brand_analysis deps required")
class XlsxIntakeTests(unittest.TestCase):
    def test_xlsx_parses_same_as_csv(self) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Line Item", "2024", "2023"])
        ws.append(["Net Revenue", 960000, 810000])
        ws.append(["COGS", 370000, 320000])
        ws.append(["Total Marketing", 250000, 210000])
        ws.append(["Gross Profit", 590000, 490000])
        ws.append(["Net Earnings", 95000, 70000])
        buf = io.BytesIO()
        wb.save(buf)
        result = intake_mod.parse_dump([("Acme_FY2024.xlsx", buf.getvalue())])
        self.assertTrue(result.has_yoy)
        self.assertEqual(result.current.net_revenue_cents, 96_000_000)
        self.assertEqual(result.current.cogs_cents, 37_000_000)


@unittest.skipUnless(DEPS_AVAILABLE, "brand_analysis deps required")
class DocxExportTests(unittest.TestCase):
    def test_docx_builds(self) -> None:
        try:
            from sales_support_agent.services.brand_analysis.docx_export import build_docx
        except ModuleNotFoundError:
            self.skipTest("python-docx not installed")
        report = build_report([_load("acme_pnl_2024.csv")], brand="Acme", use_llm=False)
        data = build_docx(report)
        self.assertGreater(len(data), 1000)
        self.assertEqual(data[:2], b"PK")  # docx is a zip container


@unittest.skipUnless(DEPS_AVAILABLE, "brand_analysis deps required")
class ParserBugRegressionTests(unittest.TestCase):
    """Regression tests for parser bugs fixed in the audit pass."""

    def _make_qbo_xlsx(self) -> bytes:
        """QBO-style P&L: revenue spread across income accounts, 'Total for Income'
        is the real total, years in column headers."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Profit and Loss"
        rows = [
            ["Profit and Loss", "", "Jan - Dec 2025", "%", "Jan - Dec 2024", "%"],
            ["Income", "", "", "", "", ""],
            ["400011 Product Sales", "", "850000", "75.9%", "900000", "78.3%"],
            ["400022 Amazon Sales",  "", "270483", "24.1%", "249567", "21.7%"],
            ["Total for Income",     "", "1120483", "100.0%", "1149567", "100.0%"],
            ["Cost of Goods Sold",   "", "", "", "", ""],
            ["500100 Product Cost",  "", "448193", "40.0%", "459827", "40.0%"],
            ["Gross Profit",         "", "672290", "60.0%", "689740", "60.0%"],
            ["Net Income",           "", "470603", "42.0%", "505809", "44.0%"],
        ]
        for r in rows: ws.append(r)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def test_qbo_prior_revenue_from_income_total_not_sub_line(self) -> None:
        """Bug: prior.net_revenue_or_derived() used to return $900K (first income
        sub-line) instead of $1,149,567 (Total for Income). This caused a fake
        +24% YoY growth (Revenue A) when the brand was actually down -2.5% (D)."""
        data = self._make_qbo_xlsx()
        result = intake_mod.parse_dump([("PnL_2025.xlsx", data)], use_llm=False)
        self.assertTrue(result.has_yoy)
        # Current revenue: _income_total picks up "Total for Income" = $1,120,483
        self.assertEqual(result.current.net_revenue_cents, 112_048_300)
        # Prior revenue: must also come from "Total for Income" col B, not $900K sub-line
        self.assertIsNotNone(result.prior)
        prior_rev = result.prior.net_revenue_or_derived()  # type: ignore[union-attr]
        self.assertEqual(prior_rev, 114_956_700)  # $1,149,567

    def test_qbo_revenue_grade_reflects_true_growth(self) -> None:
        """With correct prior revenue the revenue grade must be D (-2.5% YoY),
        not the former false A (+24% from wrong prior)."""
        from sales_support_agent.services.brand_analysis.schema import NOT_ASSESSED
        data = self._make_qbo_xlsx()
        result = intake_mod.parse_dump([("PnL_2025.xlsx", data)], use_llm=False)
        scored = scoring_mod.score(result.current, result.prior, category="dtc")
        rev_dim = next(d for d in scored["scorecard"].dimensions if d.key == "revenue")
        self.assertNotEqual(rev_dim.letter, "A")   # must NOT be fake A
        self.assertEqual(rev_dim.letter, "D")       # -2.5% YoY → D

    def test_xero_yoy_detected_from_title_row(self) -> None:
        """Bug: Xero P&L puts years in row 4 not the header, so both 2025 and
        2024 must be found via _years_of_file scanning title rows. Prior period
        was never created before because only max-year was added to all_years."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "P&L"
        rows = [
            ["Your Company", "", "", ""],
            ["Profit & Loss", "", "", ""],
            ["For the year ended 31 December 2025", "", "", ""],
            ["", "2025", "2024", ""],
            ["Revenue", "", "", ""],
            ["Sales Revenue", "1120483", "1149567", ""],
            ["Total Revenue", "1120483", "1149567", ""],
            ["Cost of Sales", "", "", ""],
            ["Purchases", "448193", "459827", ""],
            ["Gross Profit", "672290", "689740", ""],
            ["Advertising Expense", "89639", "80470", ""],
            ["Net Profit", "470603", "505809", ""],
        ]
        for r in rows: ws.append(r)
        buf = io.BytesIO(); wb.save(buf)
        result = intake_mod.parse_dump([("Xero_PnL.xlsx", buf.getvalue())], use_llm=False)
        self.assertTrue(result.has_yoy, "Xero P&L must detect prior year from title row")
        self.assertEqual(result.current.net_revenue_cents, 112_048_300)
        self.assertIsNotNone(result.prior)

    def test_media_grade_na_with_zero_spend(self) -> None:
        """Bug: _grade_media returned A when marketing_by_channel had entries but
        total spend = 0. Should return N/A — 0 spend tells us nothing."""
        from sales_support_agent.services.brand_analysis.schema import (
            PeriodFinancials, benchmarks_for, NOT_ASSESSED,
        )
        period = PeriodFinancials(
            net_revenue_cents=100_000_000,
            marketing_by_channel={"other_marketing": 0},
        )
        bm = benchmarks_for("dtc")
        letter, reason = scoring_mod._grade_media(period, bm)
        self.assertEqual(letter, NOT_ASSESSED)

    def test_media_grade_na_with_single_channel(self) -> None:
        """Bug: a single P&L 'Marketing' line created 1 channel with 100% share →
        F for media. Should be N/A — 1 channel can't show concentration."""
        from sales_support_agent.services.brand_analysis.schema import (
            PeriodFinancials, benchmarks_for, NOT_ASSESSED,
        )
        period = PeriodFinancials(
            net_revenue_cents=100_000_000,
            marketing_by_channel={"other_marketing": 89_639_00},
        )
        bm = benchmarks_for("dtc")
        letter, reason = scoring_mod._grade_media(period, bm)
        self.assertEqual(letter, NOT_ASSESSED)

    def test_media_grade_real_with_two_channels(self) -> None:
        """Media grading must still work (and not N/A) when 2+ real channels exist."""
        from sales_support_agent.services.brand_analysis.schema import (
            PeriodFinancials, benchmarks_for, NOT_ASSESSED,
        )
        period = PeriodFinancials(
            net_revenue_cents=100_000_000,
            marketing_by_channel={"meta": 15_000_000, "google": 7_000_000},
        )
        bm = benchmarks_for("dtc")
        letter, reason = scoring_mod._grade_media(period, bm)
        self.assertNotEqual(letter, NOT_ASSESSED)   # must produce a real grade
        # meta = 15M / (15+7)M = 68% share → C (>65% but ≤80%)
        self.assertEqual(letter, "C")


@unittest.skipUnless(DEPS_AVAILABLE, "brand_analysis deps required")
class MultiYearIntakeTests(unittest.TestCase):
    """Regression tests for 3+ year uploads."""

    def _make_single_year_xlsx(self, year: int, revenue: int) -> bytes:
        """Minimal single-year P&L file named for that year."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L"
        ws.append(["Line Item", str(year)])
        ws.append(["Net Revenue", revenue])
        ws.append(["COGS", revenue // 3])
        ws.append(["Net Earnings", revenue // 10])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _make_three_col_xlsx(self) -> bytes:
        """QBO-style P&L with 3 year columns: 2025 | 2024 | 2023."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profit and Loss"
        ws.append(["", "Jan - Dec 2025", "Jan - Dec 2024", "Jan - Dec 2023"])
        ws.append(["Income", "", "", ""])
        ws.append(["Product Sales", "1000000", "900000", "800000"])
        ws.append(["Total for Income", "1000000", "900000", "800000"])
        ws.append(["COGS", "400000", "360000", "320000"])
        ws.append(["Net Income", "200000", "180000", "160000"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_three_separate_files_no_contamination(self) -> None:
        """Bug: FY2023 file was routed to current (file_is_prior=False) and
        backfilled any missing current fields with 2023 data. FY2023 must be
        skipped entirely — only FY2025 (current) and FY2024 (prior) are used."""
        files = [
            ("FY2025_PnL.xlsx", self._make_single_year_xlsx(2025, 1_000_000)),
            ("FY2024_PnL.xlsx", self._make_single_year_xlsx(2024, 900_000)),
            ("FY2023_PnL.xlsx", self._make_single_year_xlsx(2023, 500_000)),  # must be ignored
        ]
        result = intake_mod.parse_dump(files, use_llm=False)
        self.assertEqual(result.current.year, 2025)
        self.assertIsNotNone(result.prior)
        self.assertEqual(result.prior.year, 2024)
        # Current revenue must be 2025's $1M, NOT contaminated by 2023's $500K
        self.assertEqual(result.current.net_revenue_cents, 100_000_000)
        # Prior revenue must be 2024's $900K
        self.assertEqual(result.prior.net_revenue_cents, 90_000_000)
        # FY2023 skipped message should appear in notes
        self.assertTrue(any("2023" in n and "Skipped" in n for n in result.notes))

    def test_three_column_pnl_uses_correct_two_years(self) -> None:
        """3-column P&L (2025/2024/2023): parser must use 2025 as current,
        2024 as prior, and ignore the 2023 column entirely."""
        files = [("PnL_3yr.xlsx", self._make_three_col_xlsx())]
        result = intake_mod.parse_dump(files, use_llm=False)
        self.assertTrue(result.has_yoy, "3-column P&L must produce a YoY comparison")
        self.assertEqual(result.current.year, 2025)
        self.assertEqual(result.prior.year, 2024)
        # Current revenue = $1,000,000 (2025 column)
        self.assertEqual(result.current.net_revenue_cents, 100_000_000)
        # Prior revenue = $900,000 (2024 column, NOT 2023 = $800K)
        self.assertEqual(result.prior.net_revenue_cents, 90_000_000)

    def test_column_years_scans_title_rows(self) -> None:
        """Bug: _column_years only checked header/rows[0]. Xero-style P&Ls put
        the year row at rows[3] — this must now be detected and not fall back
        to positional assignment."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L"
        # Years appear in row 4 (index 3), NOT in the header row
        ws.append(["Company Name", "", "", ""])          # row 0
        ws.append(["Profit & Loss", "", "", ""])         # row 1
        ws.append(["Year ended 31 Dec 2025", "", "", ""]) # row 2
        ws.append(["", "2025", "2024", "2023"])          # row 3 — year labels
        ws.append(["Revenue", "", "", ""])
        ws.append(["Sales Revenue", "1000000", "900000", "800000"])
        ws.append(["Total Revenue", "1000000", "900000", "800000"])
        ws.append(["COGS", "400000", "360000", "320000"])
        ws.append(["Net Profit", "200000", "180000", "160000"])
        buf = io.BytesIO()
        wb.save(buf)
        result = intake_mod.parse_dump([("Xero_3yr.xlsx", buf.getvalue())], use_llm=False)
        self.assertEqual(result.current.year, 2025)
        self.assertIsNotNone(result.prior)
        self.assertEqual(result.prior.year, 2024)
        # Must use 2025 column ($1M), NOT fall back to positional and land on 2023 ($800K)
        self.assertEqual(result.current.net_revenue_cents, 100_000_000)
        self.assertEqual(result.prior.net_revenue_cents, 90_000_000)


if __name__ == "__main__":
    unittest.main()
