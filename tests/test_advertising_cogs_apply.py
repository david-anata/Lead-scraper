"""Tests for COGS recognition/break-even and the template-populated apply-sheet."""

from __future__ import annotations

import io
import unittest

import openpyxl

from sales_support_agent.services.advertising import intake as I
from sales_support_agent.services.advertising import normalizers as N
from sales_support_agent.services.advertising.bulk_sheets import build_apply_sheet
from sales_support_agent.services.advertising.deliverable import _asin_scorecard
from sales_support_agent.services.advertising.schema import AdRow, Recommendation, SalesRow


_COGS_CSV = b"ASIN,COGS,FBA Fee,Referral Fee\nB07NXN4F7X,12.50,3.00,2.00\nB0CC6QGY12,9.00,,\n"


class CogsNormalizeTest(unittest.TestCase):
    def test_sniff_cogs(self):
        self.assertEqual(I.sniff_kind("cogs.csv", _COGS_CSV), I.KIND_COGS)

    def test_business_report_not_cogs(self):
        biz = b"(Child) ASIN,Sessions - Total,Units Ordered,Ordered Product Sales\nB1,900,22,$320\n"
        self.assertEqual(I.sniff_kind("biz.csv", biz), I.KIND_BUSINESS)

    def test_ads_report_not_cogs(self):
        ads = b"Campaign name,Search term,Impressions,Clicks,Total cost,Purchases,Sales,Units sold\nC,x,10,2,5,0,0,0\n"
        self.assertEqual(I.sniff_kind("st.csv", ads), I.KIND_ADS_REPORT)

    def test_normalize_sums_fees(self):
        out = N.normalize_cogs_csv(_COGS_CSV)
        self.assertEqual(out["asin"]["B07NXN4F7X"], 1750)  # 12.50 + 3 + 2
        self.assertEqual(out["asin"]["B0CC6QGY12"], 900)


class RobustnessTest(unittest.TestCase):
    """Regressions for the 500: a binary/XLSX file must never crash a CSV path."""

    def _xlsx_bytes(self, header, row):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(header)
        ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_binary_to_csv_reader_returns_empty(self):
        # A DSP report uploaded as XLSX previously raised _csv.Error here.
        dsp_xlsx = self._xlsx_bytes(["Order Name", "Total Cost", "Impressions"], ["X", 100, 5000])
        self.assertEqual(N.normalize_dsp_csv(dsp_xlsx), [])  # no raise
        self.assertEqual(N.normalize_business_report_csv(dsp_xlsx), [])

    def test_dsp_xlsx_sniffs_to_unknown_not_dsp(self):
        dsp_xlsx = self._xlsx_bytes(["Order Name", "Line Item Name", "Total Cost"], ["o", "l", 10])
        # filename says DSP, but it's a workbook -> resolved by content, not name.
        self.assertEqual(I.sniff_kind("May DSP_Zantrex.xlsx", dsp_xlsx), I.KIND_UNKNOWN)

    def test_cogs_xlsx_detected_and_parsed(self):
        cogs_xlsx = self._xlsx_bytes(["ASIN", "CoGS", "FBA Fee"], ["B0TEST", 5.36, 3.90])
        self.assertEqual(I.sniff_kind("margins.xlsx", cogs_xlsx), I.KIND_COGS)
        out = N.normalize_cogs_csv(cogs_xlsx)
        self.assertEqual(out["asin"]["B0TEST"], 926)  # 5.36 + 3.90

    def test_cogs_name_match_maps_product_to_asin(self):
        # Margin sheet keyed by product name (no ASIN) maps via Business Report title.
        margins = self._xlsx_bytes(
            ["Product Family", "Product", "Status", "CoGS", "FBA Fee", "Freight"],
            ["Zantrex - Skinnystix", "Berry, 30ct", "Base Price", -9.30, -4.16, -2.20])
        sales = [SalesRow(asin="B07NXN4F7X", title="Zantrex SkinnyStix Berry, 30 Stix", units=10,
                          ordered_product_sales_cents=350000)]
        out = N.normalize_cogs_csv(margins, sales_rows=sales)
        self.assertEqual(out["asin"]["B07NXN4F7X"], 1566)  # 9.30 + 4.16 + 2.20
        # The match is recorded so the operator can review it in the COGS Mapping tab.
        self.assertIn("name-matched", out["source"]["B07NXN4F7X"])
        self.assertIn("Skinnystix", out["source"]["B07NXN4F7X"])


class BreakEvenTest(unittest.TestCase):
    def test_break_even_and_profit_verdict(self):
        ads = [AdRow("SP", "product_ad", campaign_name="C", entity_text="B1",
                     spend_cents=2000000, sales_cents=3100000, orders=100, raw={"Advertised product ID": "B1"})]
        # price = 3,000,000c / 100 units = $300/unit; COGS $90 -> break-even ACoS 70%.
        sales = [SalesRow(asin="B1", title="Widget", units=100, sessions=5000,
                          ordered_product_sales_cents=3000000, conversion_bps=2000)]
        card = _asin_scorecard(ads, sales, 3000, {"asin": {"B1": 9000}, "sku": {}})[0]
        self.assertEqual(card["cogs_cents"], 9000)
        self.assertEqual(card["breakeven_acos_bps"], 7000)  # (300-90)/300
        # ad ACoS = 2,000,000/3,100,000 = 64.5% < 70% break-even -> profitable
        self.assertIn("Profitable", card["verdict"])

    def test_unprofitable_when_acos_over_breakeven(self):
        ads = [AdRow("SP", "product_ad", campaign_name="C", entity_text="B1",
                     spend_cents=2800000, sales_cents=3000000, orders=100, raw={"Advertised product ID": "B1"})]
        sales = [SalesRow(asin="B1", title="Widget", units=100,
                          ordered_product_sales_cents=3000000, conversion_bps=2000)]
        card = _asin_scorecard(ads, sales, 3000, {"asin": {"B1": 9000}, "sku": {}})[0]
        # ACoS 93% > 70% break-even
        self.assertIn("Unprofitable", card["verdict"])


def _rec(action, **bulk):
    bulk["action"] = action
    bulk.setdefault("ad_type", "SP")
    return Recommendation(category="x", title=action, is_bulk_actionable=True, bulk_row=bulk)


class ApplySheetTest(unittest.TestCase):
    def test_populates_template_with_ids(self):
        recs = [
            _rec("create_negative", campaign_id="111", ad_group_id="222",
                 keyword_text="cheap junk", match_type="negative exact", campaign_name="C", ad_group_name="AG"),
            _rec("create_keyword", campaign_id="333", ad_group_id="444",
                 keyword_text="widget green", match_type="exact", new_bid_cents=120, campaign_name="C", ad_group_name="AG"),
        ]
        res = build_apply_sheet(recs)
        self.assertEqual(res.applied, 2)
        wb = openpyxl.load_workbook(io.BytesIO(res.xlsx_bytes))
        ws = wb["Sponsored Products Campaigns"]
        hdr = [c.value for c in ws[1]]
        gi = lambda r, n: r[hdr.index(n)]
        data = [r for r in ws.iter_rows(min_row=2) if gi(r, "Entity").value]
        neg = next(r for r in data if gi(r, "Keyword Text").value == "cheap junk")
        self.assertEqual(gi(neg, "Entity").value, "Negative Keyword")
        self.assertEqual(gi(neg, "Operation").value, "Create")
        self.assertEqual(str(gi(neg, "Campaign ID").value), "111")
        self.assertEqual(gi(neg, "Match Type").value, "negativeExact")
        kw = next(r for r in data if gi(r, "Keyword Text").value == "widget green")
        self.assertEqual(gi(kw, "Entity").value, "Keyword")
        self.assertEqual(gi(kw, "Bid").value, 1.20)

    def test_skips_recs_without_ids(self):
        recs = [_rec("create_negative", keyword_text="x", match_type="negative exact")]
        res = build_apply_sheet(recs)
        self.assertEqual(res.applied, 0)
        self.assertEqual(res.skipped, 1)

    def test_set_bid_emits_update_row(self):
        recs = [_rec("set_bid", campaign_id="111", ad_group_id="222", keyword_id="999",
                     keyword_text="weight loss drinks", match_type="broad", new_bid_cents=57)]
        res = build_apply_sheet(recs)
        self.assertEqual(res.applied, 1)
        wb = openpyxl.load_workbook(io.BytesIO(res.xlsx_bytes))
        ws = wb["Sponsored Products Campaigns"]
        hdr = [c.value for c in ws[1]]
        gi = lambda r, n: r[hdr.index(n)]
        row = next(r for r in ws.iter_rows(min_row=2) if gi(r, "Keyword ID").value == "999")
        self.assertEqual(gi(row, "Entity").value, "Keyword")
        self.assertEqual(gi(row, "Operation").value, "Update")
        self.assertEqual(gi(row, "Bid").value, 0.57)

    def test_set_bid_skips_without_keyword_id(self):
        recs = [_rec("set_bid", campaign_id="1", ad_group_id="2", new_bid_cents=57)]  # no keyword_id
        res = build_apply_sheet(recs)
        self.assertEqual(res.applied, 0)
        self.assertEqual(res.skipped, 1)


class BulkKeywordScopeTest(unittest.TestCase):
    HEADER = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Keyword ID",
              "Campaign Name (Informational only)", "Ad Group Name (Informational only)",
              "ASIN (Informational only)", "Bid", "Keyword Text", "Match Type",
              "Impressions", "Clicks", "Spend", "Sales", "Orders", "Units"]

    def _bulk(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sponsored Products Campaigns"
        ws.append(self.HEADER)

        def row(**k):
            ws.append([k.get(h, "") for h in self.HEADER])

        # C1 = Zantrex-only
        row(Entity="Product Ad", **{"Campaign ID": "C1", "ASIN (Informational only)": "B0ZAN"})
        row(Entity="Keyword", **{"Campaign ID": "C1", "Ad Group ID": "A1", "Keyword ID": "K1",
                                 "Bid": 2.0, "Keyword Text": "kw zantrex", "Match Type": "Broad",
                                 "Impressions": 1000, "Clicks": 40, "Spend": 80, "Sales": 40, "Orders": 2, "Units": 2})
        # C2 = other brand only
        row(Entity="Product Ad", **{"Campaign ID": "C2", "ASIN (Informational only)": "B0SERO"})
        row(Entity="Keyword", **{"Campaign ID": "C2", "Ad Group ID": "A2", "Keyword ID": "K2",
                                 "Bid": 1.0, "Keyword Text": "kw sero", "Match Type": "Exact"})
        # C3 = MIXED (Zantrex + other) -> must be excluded
        row(Entity="Product Ad", **{"Campaign ID": "C3", "ASIN (Informational only)": "B0ZAN"})
        row(Entity="Product Ad", **{"Campaign ID": "C3", "ASIN (Informational only)": "B0SERO"})
        row(Entity="Keyword", **{"Campaign ID": "C3", "Ad Group ID": "A3", "Keyword ID": "K3",
                                 "Bid": 1.5, "Keyword Text": "kw mixed", "Match Type": "Broad"})
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_keeps_only_brand_only_campaign_keywords(self):
        from sales_support_agent.services.advertising.normalizers import normalize_bulk_keywords
        rows = normalize_bulk_keywords(self._bulk(), brand_asins={"B0ZAN"}, other_asins={"B0SERO"})
        texts = {r.entity_text for r in rows}
        self.assertEqual(texts, {"kw zantrex"})  # sero excluded (other brand), mixed excluded (cross-brand)
        r = rows[0]
        self.assertEqual(r.keyword_id, "K1")
        self.assertEqual(r.bid_cents, 200)
        self.assertEqual(r.spend_cents, 8000)
        self.assertEqual(r.orders, 2)


if __name__ == "__main__":
    unittest.main()
