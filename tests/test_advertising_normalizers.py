"""Tests for the Amazon CSV/XLSX normalizers (tolerant header mapping)."""

from __future__ import annotations

import io
import unittest

import openpyxl

from sales_support_agent.services.advertising import normalizers as N


def _bulk_xlsx() -> bytes:
    header = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Keyword ID",
              "Campaign Name (Informational only)", "Ad Group Name (Informational only)",
              "Keyword Text", "Match Type", "Bid", "Impressions", "Clicks", "Spend", "Sales", "Orders", "Units"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sponsored Products Campaigns"
    ws.append(header)

    def row(**k):
        ws.append([k.get(h, "") for h in header])

    row(Entity="Campaign", **{"Campaign ID": "C1", "Campaign Name (Informational only)": "Brand"})
    row(Entity="Ad Group", **{"Campaign ID": "C1", "Ad Group ID": "A1"})
    row(Entity="Keyword", **{"Campaign ID": "C1", "Ad Group ID": "A1", "Keyword ID": "K1",
                             "Campaign Name (Informational only)": "Brand",
                             "Ad Group Name (Informational only)": "AG",
                             "Keyword Text": "widget blue", "Match Type": "exact", "Bid": 1.20,
                             "Impressions": 1000, "Clicks": 40, "Spend": 40.00, "Sales": 20.00,
                             "Orders": 2, "Units": 2})
    # A structural row with no metrics should be skipped.
    row(Entity="Keyword", **{"Campaign ID": "C1", "Ad Group ID": "A1", "Keyword Text": "dead kw",
                             "Match Type": "exact", "Bid": 0.50})
    # A second sheet that isn't an ad-type sheet should be ignored.
    ws2 = wb.create_sheet("Portfolios")
    ws2.append(["Portfolio ID", "Portfolio Name"])
    ws2.append(["P1", "Main"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class BulkXlsxTest(unittest.TestCase):
    def test_parses_keyword_rows_with_metrics(self):
        rows = N.normalize_bulk_xlsx(_bulk_xlsx())
        self.assertEqual(len(rows), 1)  # only the keyword with performance
        r = rows[0]
        self.assertEqual(r.ad_type, "SP")
        self.assertEqual(r.entity_level, "keyword")
        self.assertEqual(r.entity_text, "widget blue")
        self.assertEqual(r.bid_cents, 120)
        self.assertEqual(r.spend_cents, 4000)
        self.assertEqual(r.sales_cents, 2000)
        self.assertEqual(r.orders, 2)

    def test_bad_bytes_returns_empty(self):
        self.assertEqual(N.normalize_bulk_xlsx(b"not a workbook"), [])


class SearchTermTest(unittest.TestCase):
    CSV = (
        b"Campaign Name,Ad Group Name,Customer Search Term,Match Type,Impressions,Clicks,Spend,"
        b"7 Day Total Sales,7 Day Total Orders (#),7 Day Total Units (#)\n"
        b"Brand,AG,cheap junk,exact,500,25,$50.00,0,0,0\n"
        b"Brand,AG,widget green,exact,300,15,$15.00,$60.00,4,4\n"
    )

    def test_parses_terms(self):
        rows = N.normalize_search_term_csv(self.CSV)
        self.assertEqual(len(rows), 2)
        self.assertTrue(all(r.entity_level == "search_term" for r in rows))
        green = next(r for r in rows if r.entity_text == "widget green")
        self.assertEqual(green.sales_cents, 6000)
        self.assertEqual(green.orders, 4)


class NewConsoleReportTest(unittest.TestCase):
    """The real Amazon reporting-console format: Total cost / Purchases / Sales /
    Units sold, with the entity column varying by report type."""

    SEARCH_TERM = (
        b"Budget currency,Date range,Campaign name,Ad group name,Search term,"
        b"Impressions,Clicks,CTR,Total cost,Purchases,Sales,Units sold\n"
        b"USD,\"May 07, 2026 - May 28, 2026\",Quartile Zantrex,AG_P,protein packets for water,"
        b"130,20,15%,12.10,3,90.00,3\n"
    )
    ADVERTISED = (
        b"Budget currency,Campaign name,Ad group name,Advertised product SKU,"
        b"Impressions,Clicks,Total cost,Purchases,Sales,Units sold\n"
        b"USD,Quartile Serovital,AG_P,SV_SmileCare_FBA,600,12,30.00,4,200.00,4\n"
    )
    LEGACY_ADGROUP = (
        b"State,Ad group name,Status,Default bid (USD),Keywords,Products,Impressions,"
        b"Clicks,CTR,Total cost (USD),CPC (USD),Purchases,Sales (USD),ACOS,ROAS\n"
        b"ENABLED,AG_P_B0CC,ENABLED,2,1,2,166,18,0.10,15.06,0.84,10,294.90,0.05,19.58\n"
    )

    def test_search_term_new_console(self):
        rows = N.normalize_ads_report_csv(self.SEARCH_TERM)
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r.entity_level, "search_term")
        self.assertEqual(r.entity_text, "protein packets for water")
        self.assertEqual(r.spend_cents, 1210)     # "Total cost" 12.10
        self.assertEqual(r.sales_cents, 9000)     # "Sales" 90.00
        self.assertEqual(r.orders, 3)             # "Purchases"
        self.assertEqual(r.units, 3)              # "Units sold"

    def test_advertised_product_level(self):
        rows = N.normalize_ads_report_csv(self.ADVERTISED)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].entity_level, "product_ad")
        self.assertEqual(rows[0].entity_text, "SV_SmileCare_FBA")
        self.assertEqual(rows[0].spend_cents, 3000)
        self.assertEqual(rows[0].orders, 4)

    def test_legacy_adgroup_with_usd_columns(self):
        rows = N.normalize_ads_report_csv(self.LEGACY_ADGROUP)
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r.entity_level, "ad_group")
        self.assertEqual(r.spend_cents, 1506)     # "Total cost (USD)"
        self.assertEqual(r.sales_cents, 29490)    # "Sales (USD)"
        self.assertEqual(r.bid_cents, 200)        # "Default bid (USD)" 2

    def test_search_term_backcompat_alias(self):
        # Old "Customer Search Term" + "Spend" + "7 Day Total" still parses.
        old = (
            b"Campaign Name,Customer Search Term,Impressions,Clicks,Spend,7 Day Total Sales,7 Day Total Orders (#)\n"
            b"C,cheap junk,500,25,50.00,0,0\n"
        )
        rows = N.normalize_ads_report_csv(old)
        self.assertEqual(rows[0].entity_level, "search_term")
        self.assertEqual(rows[0].spend_cents, 5000)


class BusinessReportTest(unittest.TestCase):
    CSV = (
        b"(Child) ASIN,Title,SKU,Sessions - Total,Page Views - Total,Units Ordered,"
        b"Unit Session Percentage,Featured Offer (Buy Box) Percentage,Ordered Product Sales\n"
        b"B001,Blue Widget,SKU1,\"1,200\",1500,48,4.00%,95.5%,\"$1,920.00\"\n"
    )

    def test_parses_sales(self):
        rows = N.normalize_business_report_csv(self.CSV)
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r.asin, "B001")
        self.assertEqual(r.sessions, 1200)
        self.assertEqual(r.units, 48)
        self.assertEqual(r.ordered_product_sales_cents, 192000)
        self.assertEqual(r.conversion_bps, 400)
        self.assertEqual(r.buy_box_pct_bps, 9550)


class SqpTest(unittest.TestCase):
    CSV = (
        b"Search Query,Search Query Volume,Impressions: Total Count,Impressions: ASIN Share %,"
        b"Clicks: Total Count,Clicks: ASIN Share %,Purchases: Total Count,Purchases: ASIN Share %\n"
        b"blue widget,10000,50000,12.5%,3000,18.0%,400,22.0%\n"
    )

    def test_parses_share(self):
        rows = N.normalize_sqp_csv(self.CSV)
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r.search_query, "blue widget")
        self.assertEqual(r.search_query_volume, 10000)
        self.assertEqual(r.impression_share_bps, 1250)
        self.assertEqual(r.purchase_share_bps, 2200)


class DspTest(unittest.TestCase):
    CSV = b"Campaign Name,Impressions,Clicks,Total Cost,Total Sales,Total Orders\nDSP Brand,90000,120,$300.00,$1500.00,10\n"

    def test_parses_dsp(self):
        rows = N.normalize_dsp_csv(self.CSV)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].ad_type, "DSP")
        self.assertEqual(rows[0].spend_cents, 30000)
        self.assertEqual(rows[0].sales_cents, 150000)


class ExternalCostsTest(unittest.TestCase):
    CSV = b"Channel,Amount,Note\nFacebook,$1000.00,prospecting\nInfluencer,$500.00,Q3 deal\nTikTok,250,\n"

    def test_parses_and_maps_channels(self):
        rows = N.normalize_external_costs_csv(self.CSV)
        self.assertEqual(len(rows), 3)
        by_channel = {r.channel: r for r in rows}
        self.assertEqual(by_channel["meta"].amount_cents, 100000)  # Facebook -> meta
        self.assertEqual(by_channel["influencer"].cost_type, "commission")
        self.assertEqual(by_channel["tiktok"].amount_cents, 25000)


class PreambleTest(unittest.TestCase):
    def test_skips_preamble_rows(self):
        csv = (
            b"Detail Page Sales and Traffic\nDownloaded 2026-06-04\n"
            b"(Child) ASIN,Sessions - Total,Units Ordered,Ordered Product Sales\n"
            b"B009,300,10,$100.00\n"
        )
        rows = N.normalize_business_report_csv(csv)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].asin, "B009")


def _rich_bulk_xlsx() -> bytes:
    """A bulk file with Product Targeting (auto), Product Ad (ASIN), and an SP
    home campaign for that ASIN — exercises target-ID + SB-harvest redirect."""
    header = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Keyword ID",
              "Product Targeting ID", "Product Targeting Expression", "ASIN (Informational only)",
              "Campaign Name (Informational only)", "Ad Group Name (Informational only)",
              "Keyword Text", "Match Type"]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sponsored Products Campaigns"
    ws.append(header)

    def row(**k):
        ws.append([k.get(h, "") for h in header])

    cn, an = "Fluoro5 B005GEZGSQ | SP | Auto", "Auto AG"
    row(Entity="Product Targeting", **{"Campaign ID": "C9", "Ad Group ID": "A9",
        "Product Targeting ID": "PT123", "Product Targeting Expression": "loose-match",
        "Campaign Name (Informational only)": cn, "Ad Group Name (Informational only)": an})
    # SP home for ASIN B005GEZGSQ: a campaign advertising it, with a keyword ad group.
    row(Entity="Product Ad", **{"Campaign ID": "C5", "ASIN (Informational only)": "B005GEZGSQ",
        "Campaign Name (Informational only)": "Fluoro5 | SP | Brand"})
    row(Entity="Keyword", **{"Campaign ID": "C5", "Ad Group ID": "A5", "Keyword ID": "K5",
        "Campaign Name (Informational only)": "Fluoro5 | SP | Brand",
        "Ad Group Name (Informational only)": "Brand AG", "Keyword Text": "existing", "Match Type": "exact"})
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _sb_bulk_xlsx() -> bytes:
    header = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Keyword ID",
              "Product Targeting ID", "Product Targeting Expression", "Bid", "Keyword Text",
              "Match Type", "Creative ASINs", "Landing Page ASINs",
              "Campaign Name (Informational only)", "Ad Group Name (Informational only)",
              "Impressions", "Clicks", "Spend", "Sales", "Orders", "Units"]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sponsored Brands Campaigns"
    ws.append(header)

    def row(**k):
        ws.append([k.get(h, "") for h in header])

    # brand campaign (creative ASIN B005GEZGSQ) with a keyword carrying performance.
    row(Entity="Product Collection Ad", **{"Campaign ID": "SB1", "Creative ASINs": "B005GEZGSQ, B0CTKT88VZ",
        "Campaign Name (Informational only)": "SB | Fluoro5"})
    row(Entity="Keyword", **{"Campaign ID": "SB1", "Keyword ID": "SBK1", "Bid": 1.50, "Keyword Text": "hair oil",
        "Match Type": "Exact", "Campaign Name (Informational only)": "SB | Fluoro5",
        "Impressions": 800, "Clicks": 30, "Spend": 60.0, "Sales": 200.0, "Orders": 4, "Units": 4})
    # other-brand campaign (creative ASIN B0OTHERXXX) — must be excluded in a multi-brand run.
    row(Entity="Product Collection Ad", **{"Campaign ID": "SB2", "Creative ASINs": "B0OTHERXXX",
        "Campaign Name (Informational only)": "SB | Other"})
    row(Entity="Keyword", **{"Campaign ID": "SB2", "Keyword ID": "SBK2", "Bid": 1.0, "Keyword Text": "other",
        "Match Type": "Exact", "Campaign Name (Informational only)": "SB | Other",
        "Impressions": 100, "Clicks": 5, "Spend": 5.0, "Sales": 0.0, "Orders": 0, "Units": 0})
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


class SponsoredBrandsBulkTest(unittest.TestCase):
    def test_parses_sb_keyword_rows_with_perf_and_ids(self):
        rows = N.normalize_bulk_sb(_sb_bulk_xlsx(), {"B005GEZGSQ"}, set())
        kw = [r for r in rows if r.entity_text == "hair oil"]
        self.assertTrue(kw)
        r = kw[0]
        self.assertEqual(r.ad_type, "SB")
        self.assertEqual(r.keyword_id, "SBK1")
        self.assertEqual(r.spend_cents, 6000)
        self.assertEqual(r.sales_cents, 20000)
        self.assertEqual(r.bulk_sheet, "Sponsored Brands Campaigns")

    def test_sb_cross_brand_campaign_excluded(self):
        # In a multi-brand run, the other-brand SB campaign must never appear.
        rows = N.normalize_bulk_sb(_sb_bulk_xlsx(), {"B005GEZGSQ"}, {"B0OTHERXXX"})
        self.assertTrue(any(r.keyword_id == "SBK1" for r in rows))
        self.assertFalse(any(r.keyword_id == "SBK2" for r in rows))


class MergeDuplicateEntitiesTest(unittest.TestCase):
    def _kw(self, kid, clicks, orders, sales, bid):
        from sales_support_agent.services.advertising.schema import AdRow
        return AdRow(ad_type="SP", entity_level="keyword", keyword_id=kid, entity_text="number 4",
                     clicks=clicks, orders=orders, sales_cents=sales, spend_cents=clicks * 100, bid_cents=bid)

    def test_collapses_same_keyword_id_to_richest_row(self):
        # Same keyword arriving from report + bulk with conflicting data.
        rich = self._kw("K1", clicks=213, orders=71, sales_cents=243241, bid=436)
        thin = self._kw("K1", clicks=27, orders=2, sales_cents=11399, bid=1134)
        from sales_support_agent.services.advertising.schema import AdRow
        st = AdRow(ad_type="SP", entity_level="search_term", entity_text="x", clicks=5)  # no id → passthrough
        merged, n = N.merge_duplicate_entities([rich, thin, st])
        self.assertEqual(n, 1)
        kws = [r for r in merged if r.entity_level == "keyword"]
        self.assertEqual(len(kws), 1)               # collapsed to one
        self.assertEqual(kws[0].clicks, 213)        # richest-data row wins
        self.assertIn(st, merged)                    # search term untouched

    def test_carries_bid_forward_if_winner_lacks_one(self):
        winner = self._kw("K2", clicks=50, orders=3, sales_cents=10000, bid=0)   # more clicks, no bid
        loser = self._kw("K2", clicks=10, orders=1, sales_cents=2000, bid=275)   # has the real bid
        merged, _ = N.merge_duplicate_entities([winner, loser])
        self.assertEqual(len(merged), 1)
        self.assertEqual(merged[0].clicks, 50)
        self.assertEqual(merged[0].bid_cents, 275)  # real bid carried onto the winner

    def test_distinct_keyword_ids_not_merged(self):
        a = self._kw("A", 20, 2, 5000, 100)
        b = self._kw("B", 20, 2, 5000, 100)
        merged, n = N.merge_duplicate_entities([a, b])
        self.assertEqual(n, 0)
        self.assertEqual(len(merged), 2)


class DropExistingCreatesTest(unittest.TestCase):
    def _bulk(self):
        header = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID",
                  "Keyword Text", "Match Type", "Product Targeting Expression"]
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sponsored Products Campaigns"
        ws.append(header)
        def row(**k): ws.append([k.get(h, "") for h in header])
        row(Entity="Keyword", **{"Campaign ID": "C", "Ad Group ID": "A", "Keyword Text": "no 4 shampoo", "Match Type": "Exact"})
        row(Entity="Product Targeting", **{"Campaign ID": "C", "Ad Group ID": "A", "Product Targeting Expression": 'asin="B003H85W46"'})
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def _kwrec(self, text, action="create_keyword"):
        from sales_support_agent.services.advertising.schema import Recommendation
        return Recommendation(category="x", title=text, is_bulk_actionable=True,
            bulk_row={"action": action, "ad_type": "SP", "campaign_id": "C", "ad_group_id": "A", "keyword_text": text})

    def test_drops_punctuation_variant_of_existing_keyword(self):
        # "no. 4 shampoo" (with period) collides with existing "no 4 shampoo".
        r = self._kwrec("no. 4 shampoo")
        n = N.drop_existing_creates([r], self._bulk())
        self.assertEqual(n, 1)
        self.assertFalse(r.is_bulk_actionable)

    def test_drops_existing_asin_target(self):
        r = self._kwrec("b003h85w46")  # bare ASIN -> asin="B003H85W46", already targeted
        N.drop_existing_creates([r], self._bulk())
        self.assertFalse(r.is_bulk_actionable)

    def test_keeps_genuinely_new_keyword(self):
        r = self._kwrec("brand new phrase here")
        N.drop_existing_creates([r], self._bulk())
        self.assertTrue(r.is_bulk_actionable)


class TargetingTypeEnforcementTest(unittest.TestCase):
    def _bulk(self):
        # Auto ad group (C1/A1, advertises B005GEZGSQ) + a manual KEYWORD ad group
        # (C2/A2, same ASIN) + a manual PRODUCT-TARGETING ad group (C3/A3).
        header = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Keyword ID",
                  "Product Targeting ID", "Product Targeting Expression", "Keyword Text", "Match Type",
                  "ASIN (Informational only)",
                  "Campaign Name (Informational only)", "Ad Group Name (Informational only)"]
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sponsored Products Campaigns"
        ws.append(header)
        def row(**k): ws.append([k.get(h, "") for h in header])
        row(Entity="Product Ad", **{"Campaign ID": "C1", "ASIN (Informational only)": "B005GEZGSQ"})
        row(Entity="Product Targeting", **{"Campaign ID": "C1", "Ad Group ID": "A1", "Product Targeting Expression": "loose-match"})  # auto
        row(Entity="Product Ad", **{"Campaign ID": "C2", "ASIN (Informational only)": "B005GEZGSQ"})
        row(Entity="Keyword", **{"Campaign ID": "C2", "Ad Group ID": "A2", "Keyword ID": "K", "Keyword Text": "hair", "Match Type": "exact"})
        row(Entity="Product Ad", **{"Campaign ID": "C3", "ASIN (Informational only)": "B005GEZGSQ"})
        row(Entity="Product Targeting", **{"Campaign ID": "C3", "Ad Group ID": "A3", "Product Targeting ID": "P", "Product Targeting Expression": 'asin="B0XYZ"'})
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def test_harvests_rehomed_by_type(self):
        from sales_support_agent.services.advertising.schema import Recommendation
        def rec(text):
            return Recommendation(category="new_keyword", title=text, is_bulk_actionable=True,
                bulk_row={"action": "create_keyword", "ad_type": "SP", "campaign_id": "C1",
                          "ad_group_id": "A1", "campaign_name": "Auto B005GEZGSQ", "keyword_text": text})
        kw = rec("number 4 shampoo")       # keyword harvest in an AUTO ad group
        pt = rec("b0cx67qvjk")             # ASIN harvest in an AUTO ad group
        N.enforce_targeting_type([kw, pt], self._bulk())
        # keyword harvest -> the keyword ad group; ASIN harvest -> the PT ad group.
        self.assertEqual((kw.bulk_row["campaign_id"], kw.bulk_row["ad_group_id"]), ("C2", "A2"))
        self.assertEqual((pt.bulk_row["campaign_id"], pt.bulk_row["ad_group_id"]), ("C3", "A3"))
        self.assertTrue(kw.is_bulk_actionable and pt.is_bulk_actionable)

    def test_harvest_dropped_when_no_matching_home(self):
        from sales_support_agent.services.advertising.schema import Recommendation
        pt = Recommendation(category="new_keyword", title="x", is_bulk_actionable=True,
            bulk_row={"action": "create_keyword", "ad_type": "SP", "campaign_id": "C1", "ad_group_id": "A1",
                      "campaign_name": "Auto B0NOHOME99", "keyword_text": "b0nohome99"})  # ASIN w/ no PT home
        N.enforce_targeting_type([pt], self._bulk())
        self.assertFalse(pt.is_bulk_actionable)  # no manual PT ad group → dropped, stays in burn list


class ProductTargetingAndRedirectTest(unittest.TestCase):
    def test_target_id_backfilled_for_auto_expression(self):
        from sales_support_agent.services.advertising.schema import AdRow
        idmap = N.bulk_name_id_map(_rich_bulk_xlsx())
        self.assertEqual(idmap["target"].get(("fluoro5 b005gezgsq | sp | auto", "auto ag", "loose-match")), "PT123")
        r = AdRow(ad_type="SP", entity_level="keyword", campaign_name="Fluoro5 B005GEZGSQ | SP | Auto",
                  ad_group_name="Auto AG", entity_text="loose-match")
        N.backfill_entity_ids([r], idmap)
        self.assertEqual(r.target_id, "PT123")
        self.assertEqual((r.campaign_id, r.ad_group_id), ("C9", "A9"))

    def test_asin_expanded_expression_normalized(self):
        self.assertEqual(N._norm_target_expr('asin-expanded="B0CTKT88VZ"'), 'asin="b0ctkt88vz"')

    def test_sb_harvest_redirects_to_sp_home(self):
        home = N.bulk_sp_home_by_asin(_rich_bulk_xlsx())
        self.assertIn("B005GEZGSQ", home)
        self.assertEqual(home["B005GEZGSQ"][0], "C5")

        class _Rec:  # minimal stand-in carrying a bulk_row + flag
            def __init__(self, br): self.bulk_row = br; self.is_bulk_actionable = False
        rec = _Rec({"action": "create_keyword", "campaign_name": "Fluoro5 B005GEZGSQ | SBV | EXP",
                    "keyword_text": "number four oil"})
        n = N.redirect_harvests_to_sp([rec], home)
        self.assertEqual(n, 1)
        self.assertEqual(rec.bulk_row["campaign_id"], "C5")
        self.assertEqual(rec.bulk_row["ad_group_id"], "A5")
        self.assertEqual(rec.bulk_row["ad_type"], "SP")
        self.assertTrue(rec.is_bulk_actionable)

    def test_redirect_skips_when_no_sp_home(self):
        class _Rec:
            def __init__(self, br): self.bulk_row = br; self.is_bulk_actionable = False
        rec = _Rec({"action": "create_keyword", "campaign_name": "SB | Banner B0NOHOMEXX",
                    "keyword_text": "x"})
        self.assertEqual(N.redirect_harvests_to_sp([rec], N.bulk_sp_home_by_asin(_rich_bulk_xlsx())), 0)
        self.assertNotIn("campaign_id", rec.bulk_row)  # untouched → stays in burn list only


class BackfillEntityIdsTest(unittest.TestCase):
    def test_name_id_map_and_backfill(self):
        from sales_support_agent.services.advertising.schema import AdRow
        idmap = N.bulk_name_id_map(_bulk_xlsx())
        self.assertEqual(idmap["campaign"].get("brand"), "C1")
        self.assertEqual(idmap["ad_group"].get(("brand", "ag")), "A1")
        self.assertEqual(idmap["keyword"].get(("brand", "ag", "widget blue", "exact")), "K1")

        # Report-style rows from a legacy .xlsx: names present, IDs blank.
        st = AdRow(ad_type="SP", entity_level="search_term", campaign_name="Brand",
                   ad_group_name="AG", entity_text="cheap widget")
        kw = AdRow(ad_type="SP", entity_level="keyword", campaign_name="Brand",
                   ad_group_name="AG", entity_text="widget blue", match_type="exact")
        enriched = N.backfill_entity_ids([st, kw], idmap)
        self.assertEqual(enriched, 2)
        # Harvest/negative target gets campaign + ad-group IDs (no keyword needed).
        self.assertEqual((st.campaign_id, st.ad_group_id), ("C1", "A1"))
        # An existing keyword also resolves its Keyword ID for bid changes.
        self.assertEqual(kw.keyword_id, "K1")

    def test_backfill_noop_without_map(self):
        from sales_support_agent.services.advertising.schema import AdRow
        r = AdRow(ad_type="SP", entity_level="search_term", campaign_name="X", ad_group_name="Y")
        self.assertEqual(N.backfill_entity_ids([r], {}), 0)
        self.assertEqual(r.campaign_id, "")


if __name__ == "__main__":
    unittest.main()
