"""Growth-plan workbook generator — the strategic XLSX deliverable.

Turns the audit (metrics + ranked recommendations + the normalized rows) into a
multi-tab Excel workbook modeled on the target the user validated:
  Exec Brief · Burn List · ASIN Scorecard · Campaign Actions · Negatives to Add
  · Revenue Bridge · Data Requests

Profit-true numbers (break-even ACoS) require per-ASIN COGS; until that's
supplied the Data Requests tab flags it as the #1 gap, exactly like the target.
"""

from __future__ import annotations

import io
from collections import defaultdict
from typing import Optional

from sales_support_agent.services.advertising.schema import (
    AdRow,
    Goals,
    Recommendation,
    SalesRow,
    acos_bps,
)

# --- light styling helpers --------------------------------------------------

_TITLE_FILL = "2B3644"
_HEAD_FILL = "85BBDA"
_BAND_FILL = "F4F8FB"


def _styles():
    from openpyxl.styles import Alignment, Font, PatternFill
    return {
        "title": (Font(bold=True, size=15, color="FFFFFF"), PatternFill("solid", fgColor=_TITLE_FILL)),
        "head": (Font(bold=True, size=11, color="2B3644"), PatternFill("solid", fgColor=_HEAD_FILL)),
        "wrap": Alignment(wrap_text=True, vertical="top"),
    }


def _dollars(cents: Optional[int]) -> float:
    return round((cents or 0) / 100, 2)


def _pct(bps: Optional[int]) -> Optional[float]:
    return None if bps is None else round(bps / 10000, 4)


# --- aggregations -----------------------------------------------------------


def _asin_scorecard(ad_rows, sales_rows, target_acos_bps, cogs=None) -> list[dict]:
    cogs = cogs or {}
    cogs_asin = cogs.get("asin") or {}
    cogs_sku = cogs.get("sku") or {}
    ad_by_asin: dict[str, dict] = defaultdict(lambda: {"spend": 0, "sales": 0, "orders": 0})
    for r in ad_rows:
        if r.entity_level != "product_ad":
            continue
        asin = (r.raw.get("Advertised product ID") or r.raw.get("Advertised product Id") or "").strip()
        if not asin:
            continue
        a = ad_by_asin[asin]
        a["spend"] += r.spend_cents
        a["sales"] += r.sales_cents
        a["orders"] += r.orders

    out: list[dict] = []
    for s in sales_rows:
        if not s.asin:
            continue
        ad = ad_by_asin.get(s.asin, {})
        spend = ad.get("spend", 0)
        sales = ad.get("sales", 0)
        ad_acos = acos_bps(spend, sales) if spend else None

        cogs_cents = cogs_asin.get(s.asin) or (cogs_sku.get(s.sku) if s.sku else None)
        price = round(s.ordered_product_sales_cents / s.units) if s.units else None
        breakeven_bps = None
        if cogs_cents and price and price > cogs_cents:
            breakeven_bps = round((price - cogs_cents) / price * 10000)

        out.append({
            "asin": s.asin,
            "product": s.title[:60],
            "org_sales": s.ordered_product_sales_cents,
            "units": s.units,
            "sessions": s.sessions,
            "cvr_bps": s.conversion_bps,
            "buybox_bps": s.buy_box_pct_bps,
            "ad_spend": spend,
            "ad_sales": sales,
            "ad_acos_bps": ad_acos,
            "cogs_cents": cogs_cents,
            "breakeven_acos_bps": breakeven_bps,
            "verdict": _verdict(s.conversion_bps, s.sessions, ad_acos, target_acos_bps, breakeven_bps),
        })
    out.sort(key=lambda r: r["org_sales"], reverse=True)
    return out


def _verdict(cvr_bps, sessions, ad_acos_bps, target_acos_bps, breakeven_bps=None) -> str:
    # Profit-true when COGS is known: compare ACoS to true break-even.
    if breakeven_bps is not None and ad_acos_bps is not None:
        if ad_acos_bps > breakeven_bps:
            return "Unprofitable on ads — cut/fix to break-even"
        if cvr_bps is not None and cvr_bps < 1000 and sessions >= 3000:
            return "Profitable — but fix CVR (biggest lever)"
        return "Profitable — scale"
    if cvr_bps is not None and cvr_bps < 1000 and sessions >= 3000:
        return "Fix CVR — biggest lever"
    if cvr_bps is not None and cvr_bps >= 2000:  # strong converter deserves investment
        if ad_acos_bps is None or ad_acos_bps <= target_acos_bps * 1.7:
            return "Scale — strong converter"
        return "Scale, but tighten ACoS"
    if ad_acos_bps is not None and ad_acos_bps > target_acos_bps * 1.4:
        return "Tighten ACoS — over target"
    return "Hold / monitor"


_CHANGE_CATS = ("bid_down", "bid_up", "negative_keyword", "new_keyword")


def _rec_campaign_name(rec) -> str:
    br = getattr(rec, "bulk_row", None) or {}
    name = (br.get("campaign_name") or "").strip()
    if name:
        return name
    ref = (getattr(rec, "entity_ref", "") or "")
    return ref.split("›")[0].strip() if "›" in ref else ""


def _campaign_actions(ad_rows: list[AdRow], recommendations: list, target_acos_bps: int,
                      limit: int = 20) -> list[dict]:
    """One row per campaign, showing the SPECIFIC moves the audit is making there
    (bid trims/raises, negatives, harvests) rolled up from the recommendations,
    with spend/sales/ACoS for context. Previously this only summed
    product_ad/campaign/ad_group rows — empty for accounts whose data is all
    keyword/target level — so the tab came out blank."""
    by_campaign: dict[str, dict] = defaultdict(
        lambda: {"spend": 0, "sales": 0, "bid_down": 0, "bid_up": 0,
                 "negative_keyword": 0, "new_keyword": 0})

    # Spend/sales context: sum the performance-bearing rows (keyword/target/ad)
    # by campaign. Distinct entities don't overlap, so this is the campaign total.
    for r in ad_rows:
        if r.entity_level in ("keyword", "target", "product_ad", "campaign", "ad_group"):
            c = by_campaign[r.campaign_name or "(unnamed)"]
            c["spend"] += r.spend_cents
            c["sales"] += r.sales_cents

    # The actual changes, counted per campaign.
    for rec in recommendations:
        cat = getattr(rec, "category", "")
        if cat not in _CHANGE_CATS:
            continue
        name = _rec_campaign_name(rec) or "(unnamed)"
        by_campaign[name][cat] += 1

    rows = []
    for name, agg in by_campaign.items():
        changes = sum(agg[k] for k in _CHANGE_CATS)
        if not (changes or agg["spend"] or agg["sales"]):
            continue
        a = acos_bps(agg["spend"], agg["sales"])
        rows.append({
            "campaign": name, "spend": agg["spend"], "sales": agg["sales"], "acos_bps": a,
            "changes": changes, "action": _action_str(agg), "why": _why(a, target_acos_bps, agg),
        })
    # Campaigns WITH changes first, then by spend.
    rows.sort(key=lambda r: (0 if r["changes"] else 1, -r["spend"]))
    return rows[:limit]


def _action_str(agg: dict) -> str:
    def n(c, label):
        k = agg[c]
        return f"{label}{k}" if k else ""
    parts = [p for p in (n("bid_down", "↓"), n("bid_up", "↑")) if p]
    bids = " ".join(parts) + (" bid" + ("s" if (agg["bid_down"] + agg["bid_up"]) > 1 else "")) if parts else ""
    extra = [x for x in (
        (f"+{agg['negative_keyword']} neg" if agg["negative_keyword"] else ""),
        (f"+{agg['new_keyword']} kw" if agg["new_keyword"] else ""),
    ) if x]
    out = " · ".join([p for p in [bids] + extra if p])
    return out or "Hold"


def _why(acos, target, agg: dict) -> str:
    cut = agg["bid_down"] or agg["negative_keyword"]
    grow = agg["bid_up"] or agg["new_keyword"]
    if acos is None:
        return "Spend with no attributed sales — trim / negate" if cut else "Review"
    a, t = acos / 100, (target or 0) / 100
    if cut and not grow:
        return f"{a:.0f}% ACoS vs {t:.0f}% target — trim bids / cut waste"
    if grow and not cut:
        return f"{a:.0f}% ACoS under {t:.0f}% target — scale winners"
    if cut and grow:
        return f"{a:.0f}% ACoS — trim losers, scale winners"
    return "Near target — hold"


# --- burn-list phase mapping ------------------------------------------------

_ACTION_CATEGORIES = ("negative_keyword", "bid_down", "new_keyword", "bid_up", "structure", "budget", "placement", "dayparting")

_PHASE = {
    "negative_keyword": ("1 Stop bleed", "P0", "S"),
    "bid_down": ("1 Stop bleed", "P0", "S"),
    "new_keyword": ("2 Scale winners", "P1", "S"),
    "bid_up": ("2 Scale winners", "P1", "S"),
    "structure": ("2 Scale winners", "P1", "M"),
    "external": ("3 Strategic", "P2", "M"),
    "manual": ("3 Strategic", "P1", "M"),
}


def _impact_dollars(rec: Recommendation) -> float:
    imp = rec.projected_impact or {}
    cents = imp.get("spend_saved_cents") or imp.get("sales_upside_cents") or imp.get("sales_cents") or 0
    return _dollars(cents)


# --- workbook ---------------------------------------------------------------


def build_growth_plan(
    *,
    brand: str,
    summary: dict,
    recommendations: list[Recommendation],
    ad_rows: list[AdRow],
    sales_rows: list[SalesRow],
    goals: Goals,
    narrative: str = "",
    data_window: str = "trailing ~30 days",
    cogs: Optional[dict] = None,
    has_cogs: bool = False,
) -> bytes:
    import openpyxl

    st = _styles()
    wb = openpyxl.Workbook()
    label = (brand or "Account").strip() or "Account"
    target_acos = goals.acos_target_bps or 3000

    # ---- Exec Brief ----
    ws = wb.active
    ws.title = "Exec Brief"
    _title(ws, st, f"{label.upper()} — AMAZON GROWTH PLAN")
    ws.append([f"Executive Brief · Data window: {data_window}"])
    ws.append([])
    _head(ws, st, ["Metric", "Current", "Target", "Gap"])
    rev = summary.get("total_sales_cents", 0)
    rev_target = goals.revenue_target_cents
    _row(ws, ["Revenue", _dollars(rev), _dollars(rev_target) if rev_target else "—",
              _dollars(rev_target - rev) if rev_target else "—"], money_cols=(1, 2, 3))
    # Ad spend with a TARGET = what you can spend at the goal revenue and target
    # TACoS, and a GAP = headroom still available to invest (positive) or overspend.
    cur_ad = summary.get("ad_spend_cents", 0)
    tacos_t = goals.tacos_target_bps
    target_ad = round(rev_target * tacos_t / 10000) if (rev_target and tacos_t) else None
    _row(ws, ["Ad spend", _dollars(cur_ad),
              _dollars(target_ad) if target_ad is not None else "—",
              _dollars(target_ad - cur_ad) if target_ad is not None else "—"], money_cols=(1, 2, 3))
    if target_ad is not None:
        head_now = round(rev * tacos_t / 10000) - cur_ad  # headroom at CURRENT revenue
        _row(ws, ["  ↳ spend headroom at current revenue", "", _dollars(head_now), ""], money_cols=(2,))
    _row(ws, ["External spend", _dollars(summary.get("external_spend_cents")), "", ""], money_cols=(1,))
    _row(ws, ["ACoS", _pct(summary.get("acos_bps")), _pct(target_acos), ""], pct_cols=(1, 2))
    _row(ws, ["TACoS", _pct(summary.get("tacos_bps")), _pct(goals.tacos_target_bps), ""], pct_cols=(1, 2))
    _row(ws, ["Blended TACoS (incl. off-Amazon)", _pct(summary.get("blended_tacos_bps")), "", ""], pct_cols=(1,))
    _row(ws, ["Units", summary.get("total_units", 0), goals.units_target or "—", ""])
    ws.append([])
    if narrative:
        ws.append(["Strategic read:"])
        ws.append([narrative])
        ws.cell(ws.max_row, 1).alignment = st["wrap"]
    _widths(ws, [34, 16, 16, 16])

    # ---- Burn List ----
    ws = wb.create_sheet("Burn List")
    _title(ws, st, f"{label.upper()} GROWTH BURN LIST")
    ws.append(["Sequenced for impact. P0 = do this week. Est. $ = monthly revenue gained or spend recovered."])
    ws.append([])
    _head(ws, st, ["#", "Phase", "Action", "Detail / How", "Prio", "Effort", "Est. $/mo Impact", "Status"])
    # Lead with concrete actions (negatives / bids / harvests); strategic + manual
    # notes sort to the bottom, so #1 is always a real move.
    burn_ordered = [r for r in recommendations if r.category in _ACTION_CATEGORIES] + \
                   [r for r in recommendations if r.category not in _ACTION_CATEGORIES]
    for i, rec in enumerate(burn_ordered, start=1):
        phase, prio, effort = _PHASE.get(rec.category, ("3 Strategic", "P2", "M"))
        _row(ws, [i, phase, rec.title, rec.detail or rec.rationale, prio, effort,
                  _impact_dollars(rec), "Not started"], money_cols=(6,), wrap_cols=(2, 3))
    _widths(ws, [4, 14, 46, 60, 6, 7, 16, 12])

    # ---- ASIN Scorecard ----
    ws = wb.create_sheet("ASIN Scorecard")
    _title(ws, st, f"{label.upper()} ASIN SCORECARD")
    ws.append(["Organic from Business Report; ad data from Advertised Product report. Joined on ASIN."])
    ws.append([])
    _head(ws, st, ["ASIN", "Product", "Org Sales", "Units", "Sessions", "CVR", "Buy Box",
                   "Ad Spend", "Ad Sales", "Ad ACoS", "COGS/unit", "Break-even ACoS", "Verdict / Move"])
    for r in _asin_scorecard(ad_rows, sales_rows, target_acos, cogs):
        _row(ws, [r["asin"], r["product"], _dollars(r["org_sales"]), r["units"], r["sessions"],
                  _pct(r["cvr_bps"]), _pct(r["buybox_bps"]), _dollars(r["ad_spend"]),
                  _dollars(r["ad_sales"]), _pct(r["ad_acos_bps"]),
                  _dollars(r["cogs_cents"]) if r["cogs_cents"] else "—",
                  _pct(r["breakeven_acos_bps"]) if r["breakeven_acos_bps"] is not None else "—",
                  r["verdict"]],
             money_cols=(2, 7, 8, 10), pct_cols=(5, 6, 9, 11), wrap_cols=(1, 12))
    _widths(ws, [14, 32, 11, 7, 9, 7, 7, 10, 10, 8, 10, 13, 30])

    # ---- COGS Mapping (review the auto-detect) ----
    cogs_asin = (cogs or {}).get("asin") or {}
    cogs_src = (cogs or {}).get("source") or {}
    if cogs_asin:
        ws = wb.create_sheet("COGS Mapping")
        _title(ws, st, f"{label.upper()} COGS — AUTO-MATCH REVIEW")
        ws.append(["COGS was auto-matched from your margin sheet's product descriptions to ASINs. "
                   "Review the matches; anything wrong, override with a 2-column ASIN,COGS upload."])
        ws.append([])
        _head(ws, st, ["ASIN", "Product", "COGS/unit", "Break-even ACoS", "Matched from", "Status"])
        for s in sales_rows:
            if not s.asin:
                continue
            cost = cogs_asin.get(s.asin)
            price = round(s.ordered_product_sales_cents / s.units) if s.units else None
            be = round((price - cost) / price * 10000) if (cost and price and price > cost) else None
            src = cogs_src.get(s.asin, "")
            status = "✓ exact (ASIN)" if "exact" in src else ("auto-matched — verify" if cost else "no COGS — add manually")
            _row(ws, [s.asin, s.title[:50], _dollars(cost) if cost else "—",
                      _pct(be) if be is not None else "—", src or "—", status],
                 money_cols=(2,), pct_cols=(3,), wrap_cols=(1, 4, 5))
        _widths(ws, [14, 34, 11, 14, 40, 22])

    # ---- Campaign Actions ----
    ws = wb.create_sheet("Campaign Actions")
    _title(ws, st, f"{label.upper()} CAMPAIGN ACTIONS")
    ws.append(["Every campaign we're changing, with the specific moves (↓/↑ bids, negatives, harvests)."])
    ws.append([])
    _head(ws, st, ["Campaign", "Spend", "Sales", "ACoS", "Changes", "Moves", "Why"])
    for r in _campaign_actions(ad_rows, recommendations, target_acos):
        _row(ws, [r["campaign"], _dollars(r["spend"]), _dollars(r["sales"]), _pct(r["acos_bps"]),
                  r["changes"] or "", r["action"], r["why"]], money_cols=(1, 2), pct_cols=(3,), wrap_cols=(0, 5, 6))
    _widths(ws, [46, 12, 12, 9, 9, 18, 38])

    # ---- New Campaigns (proven search terms promoted to their own campaign) ----
    promos = [r for r in recommendations if (r.bulk_row or {}).get("action") == "create_campaign"]
    if promos:
        ws = wb.create_sheet("New Campaigns")
        _title(ws, st, "NEW CAMPAIGNS — PROVEN WINNERS TO ISOLATE")
        ws.append(["⚠️ Apply-ready ones are CREATED LIVE when you upload the Additions file. Review each before uploading."])
        ws.append([])
        _head(ws, st, ["Campaign", "Keyword (exact)", "Target SKU(s)", "Bid", "Daily budget", "State", "Why", "In Additions file?"])
        for rec in promos:
            br = rec.bulk_row or {}
            skus = ", ".join(p.get("sku", "") for p in br.get("products", [])) or "—"
            in_file = ("Yes — LIVE on upload" if rec.is_bulk_actionable
                       else f"Review only ({br.get('review_only_reason', 'unresolved')})")
            _row(ws, [br.get("campaign_name", ""), br.get("keyword_text", ""), skus,
                      _dollars(br.get("new_bid_cents")), _dollars(br.get("daily_budget_cents")),
                      br.get("state", "enabled"), rec.detail, in_file],
                 money_cols=(3, 4), wrap_cols=(0, 2, 6, 7))
        _widths(ws, [34, 24, 22, 9, 12, 10, 40, 22])

    # ---- Negatives to Add ----
    ws = wb.create_sheet("Negatives to Add")
    _title(ws, st, "NEGATIVE KEYWORDS / TARGETS TO ADD")
    ws.append(["From the Search Term report — spend with zero or unprofitable sales."])
    ws.append([])
    _head(ws, st, ["Search term / target", "Wasted spend", "Match", "Add to campaign"])
    for rec in recommendations:
        if rec.category != "negative_keyword":
            continue
        br = rec.bulk_row or {}
        _row(ws, [br.get("keyword_text", rec.entity_ref), _impact_dollars(rec),
                  "Neg exact", br.get("campaign_name", "")], money_cols=(1,), wrap_cols=(0, 3))
    _widths(ws, [40, 14, 12, 40])

    # ---- Revenue Bridge ----
    ws = wb.create_sheet("Revenue Bridge")
    _title(ws, st, f"REVENUE BRIDGE TO {(_dollars(rev_target) if rev_target else 'TARGET')}")
    ws.append(["Editable planning model — adjust the assumptions as results come in."])
    ws.append([])
    _head(ws, st, ["Input", "Value", "Source"])
    _row(ws, ["Current monthly revenue", _dollars(rev), "Business Report"], money_cols=(1,))
    _row(ws, ["Current monthly ad spend", _dollars(summary.get("ad_spend_cents")), "Advertised Product"], money_cols=(1,))
    _row(ws, ["Target monthly revenue", _dollars(rev_target) if rev_target else "—", "Goal"], money_cols=(1,))
    gap = (rev_target - rev) if rev_target else 0
    _row(ws, ["Revenue gap to close", _dollars(gap) if rev_target else "—", "Computed"], money_cols=(1,))
    ws.append([])
    ws.append(["Levers (planning estimates):"])
    _row(ws, ["Recover wasted spend (negatives)", _sum_impact(recommendations, "negative_keyword"), "Burn List P0"], money_cols=(1,))
    _row(ws, ["Scale winners (bid-up + harvest)", _sum_impact(recommendations, "bid_up", "new_keyword"), "Burn List P1"], money_cols=(1,))
    _widths(ws, [38, 16, 22])

    # ---- Data Requests ----
    ws = wb.create_sheet("Data Requests")
    _title(ws, st, "REPORTS NEEDED TO MAKE THIS PROFIT-TRUE")
    ws.append(["What's missing, why it matters, and priority. COGS is the #1 gap for guaranteeing profitability."])
    ws.append([])
    _head(ws, st, ["#", "Report / data", "Where / how to get it", "Why it matters", "Priority"])
    reqs = []
    if not has_cogs:
        reqs.append((
            "Per-ASIN COGS + FBA & referral fees",
            "Your own cost data: one row per ASIN with landed unit cost (COGS + FBA fee + referral fee + freight). Upload as a 2-column ASIN,COGS CSV (extra fee columns optional).",
            "The #1 gap. Enables TRUE break-even ACoS per product — guarantee (not estimate) each SKU stays profitable. Today ACoS/TACoS are proxies.",
            "P0", ""))
    reqs += [
        ("Targeting / Keyword report",
         "Amazon Ads Console → Reports → run the 'Targeting' template (Sponsored Products) → Use template → export CSV.",
         "Unlocks keyword-level bid moves (retune existing bids, not just harvest/negate) — adds bid-change rows to the apply sheet.", "P0",
         "https://advertising.amazon.com/reports"),
        ("FBA Inventory & Days-of-Supply",
         "Seller Central → Reports → Fulfillment → 'FBA Inventory' / 'Restock Inventory' → download.",
         "Can't scale if best-sellers stock out — sets reorder triggers + caps spend on low-stock SKUs.", "P1",
         "https://sellercentral.amazon.com/reportcentral"),
        ("Placement report (TOS / Product / Rest)",
         "Amazon Ads Console → Reports → 'Placement' template → export CSV.",
         "Unlocks placement bid modifiers — a fast 5–15% ACoS win by over-weighting Top-of-Search on high-CVR ASINs.", "P1",
         "https://advertising.amazon.com/reports"),
    ]
    for i, (name, where, why, prio, url) in enumerate(reqs, start=1):
        _row(ws, [i, name, where, why, prio], wrap_cols=(1, 2, 3))
        if url:
            cell = ws.cell(ws.max_row, 3)  # the 'Where' cell -> clickable
            cell.hyperlink = url
            cell.style = "Hyperlink"
    _widths(ws, [4, 30, 50, 46, 8])

    _apply_conditional_formatting(wb)

    # Final pass: stretch the dark title banner (row 1) across each sheet's width.
    title_font, title_fill = st["title"]
    for sheet in wb.worksheets:
        for col in range(1, (sheet.max_column or 1) + 1):
            c = sheet.cell(1, col)
            c.fill = title_fill
            if col == 1:
                c.font = title_font
        sheet.row_dimensions[1].height = 26

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sum_impact(recs: list[Recommendation], *categories: str) -> float:
    total = 0
    for r in recs:
        if r.category in categories:
            imp = r.projected_impact or {}
            total += imp.get("spend_saved_cents") or imp.get("sales_upside_cents") or imp.get("sales_cents") or 0
    return _dollars(total)


# --- low-level sheet writers ------------------------------------------------


def _title(ws, st, text: str) -> None:
    ws.append([text])
    cell = ws.cell(ws.max_row, 1)
    cell.font, cell.fill = st["title"]


def _head(ws, st, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Border, Side
    thin = Side(style="thin", color="C8D3DD")
    ws.append(headers)
    r = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(r, col)
        cell.font, cell.fill = st["head"]
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[r].height = 26
    ws.freeze_panes = f"A{r + 1}"  # keep the header visible; string form avoids creating an empty row
    ws._adv_cols = len(headers)          # table width, for banding + borders
    ws._adv_band = 0


def _row(ws, values, *, money_cols=(), pct_cols=(), wrap_cols=()) -> None:
    from openpyxl.styles import Alignment, Border, PatternFill, Side
    ws.append(values)
    r = ws.max_row
    # money_cols / pct_cols / wrap_cols are 0-based indices into `values`.
    for c in money_cols:
        cell = ws.cell(r, c + 1)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "$#,##0.00"
    for c in pct_cols:
        cell = ws.cell(r, c + 1)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.0%"

    wrap_set = {c + 1 for c in wrap_cols}
    cols = getattr(ws, "_adv_cols", len(values))
    ws._adv_band = getattr(ws, "_adv_band", 0) + 1
    fill = PatternFill("solid", fgColor=_BAND_FILL) if ws._adv_band % 2 == 0 else None
    thin = Side(style="thin", color="EAEEF2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(r, col)
        cell.border = border
        if fill is not None:
            cell.fill = fill
        cell.alignment = Alignment(wrap_text=col in wrap_set, vertical="top")


def _widths(ws, widths: list[int]) -> None:
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --- conditional formatting (color scales, data bars, highlight rules) ------

_GREEN = ("C6EFCE", "006100")
_RED = ("FFC7CE", "9C0006")
_AMBER = ("FFEB9C", "9C6500")


def _apply_conditional_formatting(wb) -> None:
    """Make every table scannable: color scales on rate columns, data bars on $
    impact, and text/threshold highlights on verdicts and statuses."""
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule, Rule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle

    def _scale(ws, rng, lo, hi):  # lo=color at min value, hi=color at max value
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color=lo, mid_type="percentile", mid_value=50,
            mid_color="FFEB84", end_type="max", end_color=hi))

    def _bar(ws, rng, color):
        ws.conditional_formatting.add(rng, DataBarRule(start_type="min", end_type="max", color=color))

    def _contains(ws, rng, text, pair):
        fill, font = pair
        dxf = DifferentialStyle(fill=PatternFill(bgColor=fill), font=Font(color=font))
        rule = Rule(type="containsText", operator="containsText", text=text, dxf=dxf)
        rule.formula = [f'NOT(ISERROR(SEARCH("{text}",{rng.split(":")[0]})))']
        ws.conditional_formatting.add(rng, rule)

    def _formula(ws, rng, formula, pair):
        fill, font = pair
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[formula], fill=PatternFill(bgColor=fill), font=Font(color=font)))

    def _range(ws, first_col_value):
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and row[0] == first_col_value:
                return i + 1, ws.max_row
        return None, None

    if "ASIN Scorecard" in wb.sheetnames:
        ws = wb["ASIN Scorecard"]; f, l = _range(ws, "ASIN")
        if f and l >= f:
            _scale(ws, f"J{f}:J{l}", "63BE7B", "F8696B")   # Ad ACoS: low=green, high=red
            _scale(ws, f"F{f}:F{l}", "F8696B", "63BE7B")   # CVR: low=red, high=green
            _contains(ws, f"M{f}:M{l}", "Scale", _GREEN)
            _contains(ws, f"M{f}:M{l}", "Unprofitable", _RED)
            _contains(ws, f"M{f}:M{l}", "Tighten", _AMBER)
            _contains(ws, f"M{f}:M{l}", "Fix CVR", _AMBER)

    if "Burn List" in wb.sheetnames:
        ws = wb["Burn List"]; f, l = _range(ws, "#")
        if f and l >= f:
            _bar(ws, f"G{f}:G{l}", "63BE7B")               # $ impact bars
            _contains(ws, f"E{f}:E{l}", "P0", _RED)
            _contains(ws, f"E{f}:E{l}", "P1", _AMBER)

    if "Campaign Actions" in wb.sheetnames:
        ws = wb["Campaign Actions"]; f, l = _range(ws, "Campaign")
        if f and l >= f:
            _scale(ws, f"D{f}:D{l}", "63BE7B", "F8696B")   # ACoS
            _contains(ws, f"E{f}:E{l}", "PAUSE", _RED)
            _contains(ws, f"E{f}:E{l}", "CUT", _AMBER)
            _contains(ws, f"E{f}:E{l}", "SCALE", _GREEN)

    if "Negatives to Add" in wb.sheetnames:
        ws = wb["Negatives to Add"]; f, l = _range(ws, "Search term / target")
        if f and l >= f:
            _bar(ws, f"B{f}:B{l}", "F8696B")               # wasted spend (red bars)

    if "COGS Mapping" in wb.sheetnames:
        ws = wb["COGS Mapping"]; f, l = _range(ws, "ASIN")
        if f and l >= f:
            _contains(ws, f"F{f}:F{l}", "exact", _GREEN)
            _contains(ws, f"F{f}:F{l}", "verify", _AMBER)
            _contains(ws, f"F{f}:F{l}", "no COGS", _RED)

    if "Exec Brief" in wb.sheetnames:
        ws = wb["Exec Brief"]
        rows = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and isinstance(row[0], str):
                rows.setdefault(row[0].strip(), i)
        if "Revenue" in rows:  # positive revenue gap = behind goal = red
            _formula(ws, f"D{rows['Revenue']}", f"=$D${rows['Revenue']}>0", _RED)
        if "Ad spend" in rows:  # positive ad-spend gap = headroom to invest = green
            _formula(ws, f"D{rows['Ad spend']}", f"=$D${rows['Ad spend']}>0", _GREEN)
        for metric in ("ACoS", "TACoS"):
            if metric in rows:  # current worse (higher) than target = red
                r = rows[metric]
                _formula(ws, f"B{r}", f"=AND(ISNUMBER($C${r}),$B${r}>$C${r})", _RED)
