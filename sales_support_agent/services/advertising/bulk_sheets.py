"""Amazon Ads bulk-sheet generation by ROUND-TRIPPING the user's own upload.

The safest way to emit a Seller-Central-valid bulk file is to take the bulk-
operations workbook the operator already downloaded and mutate it in place:
  * set_bid        -> find the matching keyword/target row, set its Bid +
                      Operation=update
  * create_negative-> append a Negative Keyword row (Operation=create), copying
                      Campaign/Ad Group IDs from a sibling row in that ad group
  * create_keyword -> append a Keyword row (Operation=create, exact match, bid)

Untouched rows keep a blank Operation, so Amazon ignores them. This guarantees
the exact per-ad-type column schema without us reconstructing it from memory.
Ad types with no uploaded bulk file (or unsupported, e.g. STV/DSP) stay manual
tasks in the burn list.

Returns BulkBuildResult: the new xlsx bytes plus per-rec applied/skipped status,
so the page can tell the operator exactly what made it into the sheet.
"""

from __future__ import annotations

import io
import logging
import os
import re
from dataclasses import dataclass, field
from typing import Optional

from sales_support_agent.services.advertising.normalizers import _ad_type_from_sheet, _norm_key
from sales_support_agent.services.advertising.schema import (
    BULK_SUPPORTED,
    Recommendation,
)

logger = logging.getLogger(__name__)

_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "templates", "amazon_bulk_template.xlsx")

# rec match-type text -> Amazon bulk value (from the template's Config sheet)
_MATCH_MAP = {
    "negative exact": "negativeExact", "negativeexact": "negativeExact",
    "negative phrase": "negativePhrase", "negativephrase": "negativePhrase",
    "exact": "exact", "phrase": "phrase", "broad": "broad",
}

_BARE_ASIN_RE = re.compile(r"^b0[a-z0-9]{8}$", re.I)
_TARGET_EXPR_RE = re.compile(r"^\s*(asin|asin-expanded|category|brand)\s*=", re.I)
# Amazon keyword text allows letters, digits, spaces, and . - & ' (and accents).
_VALID_KEYWORD_RE = re.compile(r"^[\w .\-&']+$", re.UNICODE)


def _classify_target_text(text: str):
    """Decide how a 'search term / keyword' string must be applied:
      ('keyword', text)         -> a real keyword
      ('target', 'asin="B0…"')  -> an ASIN/category PRODUCT TARGET, not a keyword
      ('invalid', None)         -> can't be applied (bad characters)
    ASIN search terms come from auto / product-targeting placements; Amazon
    rejects them as keywords, so they must become product-targeting rows."""
    t = (text or "").strip()
    if not t:
        return ("invalid", None)
    if _TARGET_EXPR_RE.match(t):
        # normalize asin-expanded="B0…" -> asin="B0…"; uppercase the ASIN.
        expr = re.sub(r"asin-expanded\s*=", "asin=", t, flags=re.I)
        expr = re.sub(r'(asin\s*=\s*")([^"]+)(")', lambda m: m.group(1) + m.group(2).upper() + m.group(3), expr, flags=re.I)
        return ("target", expr)
    if _BARE_ASIN_RE.match(t):
        return ("target", f'asin="{t.upper()}"')
    if not _VALID_KEYWORD_RE.match(t) or len(t.split()) > 10:
        return ("invalid", None)
    return ("keyword", t)


@dataclass
class BulkBuildResult:
    xlsx_bytes: Optional[bytes] = None
    applied: int = 0
    skipped: int = 0
    applied_titles: list[str] = field(default_factory=list)
    skipped_titles: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    invalid: int = 0
    issues: list[str] = field(default_factory=list)

    @property
    def has_file(self) -> bool:
        return bool(self.xlsx_bytes) and self.applied > 0


_VALID_STATES = {"enabled", "paused", "archived"}
_VALID_MATCH_TYPES = {"exact", "phrase", "broad", "negativeExact", "negativePhrase"}
_BID_MIN, _BID_MAX = 0.02, 1000.0  # USD (Amazon SP/SB/SD)


def validate_bulk_rows(wb) -> list[str]:
    """Pre-flight: check every populated row in a built workbook against Amazon's
    rules BEFORE the file is offered for download, so review errors are caught
    in-app. Validates value constraints (State, Match Type, Bid range, keyword
    text) for all rows, and per-(entity, operation) required/optional headers for
    Sponsored Products using the template's own `Config` sheet (the 2.0 spec).
    Returns a list of human-readable issues (empty = clean)."""
    rules: dict = {}
    if "Config" in wb.sheetnames:
        for row in wb["Config"].iter_rows(values_only=True):
            cells = [c for c in row if c not in (None, "")]
            if cells:
                rules[str(cells[0])] = [str(x) for x in cells[1:]]

    def norm(h: str) -> str:
        return str(h).strip().lower()

    issues: list[str] = []
    for ws in wb.worksheets:
        if ws.title == "Config":
            continue
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        # Duplicate-ID check: Amazon rejects the whole file if an ID repeats
        # within a sheet (e.g. two Update rows for the same Keyword ID).
        for id_col in ("Keyword ID", "Product Targeting ID", "Ad ID"):
            if id_col not in hdr:
                continue
            ci = hdr.index(id_col)
            seen_ids: dict = {}
            for ri, r in enumerate(rows[1:], start=2):
                if ci >= len(r) or r[ci] in (None, ""):
                    continue
                key = str(r[ci])
                if key in seen_ids:
                    issues.append(f"{ws.title} r{ri}: duplicate {id_col} {key!r} (also row {seen_ids[key]})")
                else:
                    seen_ids[key] = ri
        for ri, r in enumerate(rows[1:], start=2):
            cell = {hdr[i]: r[i] for i in range(len(hdr)) if i < len(r)}
            entity, op, product = cell.get("Entity"), cell.get("Operation"), cell.get("Product")
            if not entity:
                continue
            tag = f"{ws.title} r{ri}"
            # --- value constraints (every product) ---
            st = cell.get("State")
            if st and str(st) not in _VALID_STATES:
                issues.append(f"{tag}: invalid State {st!r}")
            mt = cell.get("Match Type")
            if mt not in (None, "") and str(mt) not in _VALID_MATCH_TYPES:
                issues.append(f"{tag}: invalid Match Type {mt!r}")
            bid = cell.get("Bid")
            if bid not in (None, ""):
                try:
                    b = float(bid)
                    if b < _BID_MIN or b > _BID_MAX:
                        issues.append(f"{tag}: Bid {b} outside ${_BID_MIN}–${_BID_MAX}")
                except (TypeError, ValueError):
                    issues.append(f"{tag}: non-numeric Bid {bid!r}")
            kt = cell.get("Keyword Text")
            if entity in ("Keyword", "Negative Keyword") and kt:
                if _classify_target_text(str(kt))[0] != "keyword":
                    issues.append(f"{tag}: {kt!r} is not valid keyword text (ASIN/illegal char)")
            # --- per-operation headers (Sponsored Products; Config has SP rules) ---
            if product == "Sponsored Products" and op:
                key = "SponsoredProducts" + str(op) + str(entity).replace(" ", "")
                req = rules.get(key + "RequiredHeaders")
                if req is not None:
                    req_n = {norm(h) for h in req}
                    allowed = req_n | {norm(h) for h in rules.get(key + "OptionalHeaders", [])} | {"product", "entity", "operation"}
                    present = {norm(h) for h, v in cell.items() if v not in (None, "")}
                    missing = req_n - present
                    extra = present - allowed
                    if missing:
                        issues.append(f"{tag}: {op} {entity} missing required {sorted(missing)}")
                    if extra:
                        issues.append(f"{tag}: {op} {entity} has forbidden columns {sorted(extra)}")
    return issues


def build_apply_sheet(recommendations: list[Recommendation]) -> BulkBuildResult:
    """Populate Amazon's official bulk template with create-negative and
    create-keyword rows, using the Campaign ID / Ad Group ID carried on each
    recommendation (extracted from the uploaded report CSVs). Produces an
    upload-ready Sponsored Products bulk sheet WITHOUT the operator having to
    download or edit anything.

    Bid changes on existing keywords (set_bid) become Operation=update rows keyed
    by Keyword ID — available when an Amazon Bulk Operations file was uploaded.
    """
    result = BulkBuildResult()
    actionable = [
        r for r in recommendations
        if r.is_bulk_actionable and r.bulk_row.get("action") in ("create_negative", "create_keyword", "set_bid")
        and r.bulk_row.get("ad_type") in ("SP", "SB")
    ]
    if not actionable:
        result.notes.append("No create-keyword / create-negative / bid-change actions to apply.")
        return result

    try:
        import openpyxl
        wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    except Exception:
        logger.exception("[advertising] could not open bundled Amazon bulk template")
        result.notes.append("Amazon bulk template unavailable.")
        return result

    _PRODUCT = {"SP": "Sponsored Products", "SB": "Sponsored Brands"}
    _ctx_cache: dict = {}

    def _resolve_sheet(br):
        """The template sheet to write this rec into. SB entities may belong to
        'Sponsored Brands Campaigns' or 'SB Multi Ad Group Campaigns' — Amazon
        validates per sheet, so honour the source sheet when known."""
        adt = br.get("ad_type")
        want = br.get("bulk_sheet") or ""
        if adt == "SB" and want:
            for sheet in wb.worksheets:
                if sheet.title == want:
                    return sheet
        for sheet in wb.worksheets:
            if _ad_type_from_sheet(sheet.title) == adt:
                return sheet
        return None

    def _ctx(ws):
        if ws.title not in _ctx_cache:
            header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            col = {name: _col_index(header, name) for name in
                   ("Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Keyword ID",
                    "Product Targeting ID", "Product Targeting Expression",
                    "Campaign Name", "Ad Group Name", "Keyword Text", "Match Type", "Bid", "State")}
            _ctx_cache[ws.title] = (header, col)
        return _ctx_cache[ws.title]

    # Amazon rejects the WHOLE file if any ID appears twice (e.g. two Update rows
    # for the same Keyword ID). Track each row's unique key and skip repeats —
    # recs are ranked, so the first (highest-priority) one wins.
    _seen: set = set()

    for rec in actionable:
        br = rec.bulk_row
        action = br["action"]
        ws = _resolve_sheet(br)
        if ws is None:
            result.skipped += 1
            result.skipped_titles.append(rec.title)
            continue
        header, col = _ctx(ws)
        product = _PRODUCT.get(br.get("ad_type"), "Sponsored Products")
        row = [""] * len(header)

        def put(name, value, _row=row, _col=col):
            i = _col.get(name)
            if i is not None:
                _row[i] = value

        # Amazon validates each row against the operation's required/optional
        # headers (see the template's Config sheet). Writing ANY extra field
        # (informational names, or Keyword Text/Match Type on an update — which
        # can't change) fails review, so emit ONLY the allowed columns per op.
        if action == "set_bid":
            new_bid = br.get("new_bid_cents")
            if br.get("keyword_id") and new_bid:
                # Update Keyword → required: Keyword Id, State; optional: Bid.
                dup_key = ("kw_update", str(br["keyword_id"]))
                put("Product", product)
                put("Entity", "Keyword")
                put("Operation", "Update")
                put("Keyword ID", br["keyword_id"])
                put("State", "enabled")
                put("Bid", _dollars(new_bid))
            elif br.get("target_id") and new_bid:
                # Update Product Targeting → required: Product Targeting Id, State; optional: Bid.
                dup_key = ("pt_update", str(br["target_id"]))
                put("Product", product)
                put("Entity", "Product Targeting")
                put("Operation", "Update")
                put("Product Targeting ID", br["target_id"])
                put("State", "enabled")
                put("Bid", _dollars(new_bid))
            else:
                result.skipped += 1
                result.skipped_titles.append(rec.title)
                continue
            if dup_key in _seen:  # same entity already has an update row
                result.skipped += 1
                result.skipped_titles.append(rec.title)
                continue
            _seen.add(dup_key)
            ws.append(row)
            result.applied += 1
            result.applied_titles.append(rec.title)
            continue

        # create_negative / create_keyword
        if not br.get("campaign_id") or not br.get("ad_group_id"):
            result.skipped += 1
            result.skipped_titles.append(rec.title)
            continue
        is_neg = action == "create_negative"
        kind, value = _classify_target_text(br.get("keyword_text", ""))
        if kind == "invalid":
            # e.g. "#4 hair care" — not a valid keyword and not a product target.
            result.skipped += 1
            result.skipped_titles.append(rec.title)
            continue
        if kind == "target":
            # An ASIN/category search term is a PRODUCT TARGET, not a keyword —
            # Amazon rejects it as a keyword. Required: Campaign Id, Ad Group Id,
            # State, Product Targeting Expression; optional Bid (positive only).
            dup_key = ("npt" if is_neg else "pt", str(br["campaign_id"]), str(br["ad_group_id"]), value.lower())
        else:
            match_type = _MATCH_MAP.get(_norm_key(br.get("match_type")), "negativeExact" if is_neg else "exact")
            dup_key = ("neg" if is_neg else "kw", str(br["campaign_id"]), str(br["ad_group_id"]), value.lower(), match_type)
        if dup_key in _seen:  # same create already emitted (dup keyword/target in the ad group)
            result.skipped += 1
            result.skipped_titles.append(rec.title)
            continue
        _seen.add(dup_key)
        put("Product", product)
        put("Operation", "Create")
        put("Campaign ID", br["campaign_id"])
        put("Ad Group ID", br["ad_group_id"])
        put("State", "enabled")
        if kind == "target":
            put("Entity", "Negative Product Targeting" if is_neg else "Product Targeting")
            put("Product Targeting Expression", value)
            if not is_neg and br.get("new_bid_cents"):
                put("Bid", _dollars(br["new_bid_cents"]))
        else:
            # Create (Negative) Keyword → + Keyword Text, Match Type; optional Bid.
            put("Entity", "Negative Keyword" if is_neg else "Keyword")
            put("Keyword Text", value)
            put("Match Type", match_type)
            if not is_neg and br.get("new_bid_cents"):
                put("Bid", _dollars(br["new_bid_cents"]))
        ws.append(row)
        result.applied += 1
        result.applied_titles.append(rec.title)

    if result.applied:
        # Pre-flight: validate every row against Amazon's rules (Config sheet +
        # value limits) before the file can be downloaded. Should be 0 — a
        # regression guard so a future change can't ship a review-failing sheet.
        result.issues = validate_bulk_rows(wb)
        result.invalid = len(result.issues)
        if result.issues:
            logger.warning("[advertising] apply-sheet pre-flight found %d issue(s): %s",
                           result.invalid, result.issues[:8])
            result.notes.append(
                f"⚠ Pre-flight validation flagged {result.invalid} row issue(s) before upload — "
                "see the run log; these would likely fail Amazon's review."
            )
        buf = io.BytesIO()
        wb.save(buf)
        result.xlsx_bytes = buf.getvalue()
        result.notes.append(
            f"{result.applied} change(s) written into Amazon's bulk template — upload directly to "
            "Ads Console → Bulk operations → Upload. No manual editing needed."
        )
    return result


def _col_index(header: list[str], *aliases: str) -> Optional[int]:
    norm = [_norm_key(h) for h in header]
    for alias in aliases:
        target = _norm_key(alias)
        if target in norm:
            return norm.index(target)
    for alias in aliases:
        target = _norm_key(alias)
        for i, h in enumerate(norm):
            if target and target in h:
                return i
    return None


def _dollars(cents: int) -> float:
    return round(cents / 100, 2)


def build_bulk_workbook(uploaded_xlsx_bytes: bytes, recommendations: list[Recommendation]) -> BulkBuildResult:
    """Apply bulk-actionable recommendations to the uploaded workbook."""
    result = BulkBuildResult()

    actionable = [r for r in recommendations if r.is_bulk_actionable and r.bulk_row]
    if not actionable:
        result.notes.append("No bulk-actionable recommendations in this audit.")
        return result

    if not uploaded_xlsx_bytes:
        result.skipped = len(actionable)
        result.skipped_titles = [r.title for r in actionable]
        result.notes.append("No bulk-operations file was uploaded, so changes can't be round-tripped into a sheet.")
        return result

    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        result.notes.append("openpyxl unavailable; cannot generate bulk sheet.")
        return result

    try:
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_xlsx_bytes), data_only=False)
    except Exception:
        logger.exception("[advertising] failed to load uploaded bulk workbook")
        result.notes.append("Uploaded file could not be opened as an Amazon bulk workbook.")
        result.skipped = len(actionable)
        result.skipped_titles = [r.title for r in actionable]
        return result

    # Map each supported ad type to its sheet.
    sheet_for_type: dict[str, object] = {}
    for sheet in wb.worksheets:
        at = _ad_type_from_sheet(sheet.title)
        if at in BULK_SUPPORTED and at not in sheet_for_type:
            sheet_for_type[at] = sheet

    for rec in actionable:
        ad_type = rec.bulk_row.get("ad_type", "")
        sheet = sheet_for_type.get(ad_type)
        if sheet is None:
            result.skipped += 1
            result.skipped_titles.append(rec.title)
            continue
        ok = _apply_rec_to_sheet(sheet, rec)
        if ok:
            result.applied += 1
            result.applied_titles.append(rec.title)
        else:
            result.skipped += 1
            result.skipped_titles.append(rec.title)

    if result.applied:
        buf = io.BytesIO()
        wb.save(buf)
        result.xlsx_bytes = buf.getvalue()
    if result.skipped:
        result.notes.append(
            f"{result.skipped} change(s) could not be matched to a row in the uploaded sheet "
            "(campaign/ad-group/keyword names didn't line up) and remain manual tasks."
        )
    return result


def _sheet_header(sheet) -> list[str]:
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(c).strip() if c is not None else "" for c in row]
    return []


def _apply_rec_to_sheet(sheet, rec: Recommendation) -> bool:
    header = _sheet_header(sheet)
    if not header:
        return False

    idx = {
        "entity": _col_index(header, "Entity"),
        "operation": _col_index(header, "Operation"),
        "campaign_name": _col_index(header, "Campaign Name (Informational only)", "Campaign Name", "Campaign"),
        "ad_group_name": _col_index(header, "Ad Group Name (Informational only)", "Ad Group Name", "Ad Group"),
        "campaign_id": _col_index(header, "Campaign ID"),
        "ad_group_id": _col_index(header, "Ad Group ID"),
        "keyword_text": _col_index(header, "Keyword Text", "Product Targeting Expression"),
        "match_type": _col_index(header, "Match Type"),
        "bid": _col_index(header, "Bid"),
        "product": _col_index(header, "Product"),
    }
    if idx["operation"] is None or idx["entity"] is None:
        return False

    br = rec.bulk_row
    action = br.get("action")
    want_campaign = _norm_key(br.get("campaign_name"))
    want_ad_group = _norm_key(br.get("ad_group_name"))
    want_kw = _norm_key(br.get("keyword_text"))

    if action == "set_bid":
        return _set_bid(sheet, idx, want_campaign, want_ad_group, want_kw, br.get("new_bid_cents"))
    if action in ("create_negative", "create_keyword"):
        return _append_row(sheet, header, idx, rec)
    return False


def _cell_norm(row_cells, col: Optional[int]) -> str:
    if col is None or col >= len(row_cells):
        return ""
    v = row_cells[col]
    return _norm_key(v)


def _set_bid(sheet, idx, want_campaign, want_ad_group, want_kw, new_bid_cents) -> bool:
    if idx["bid"] is None or new_bid_cents is None:
        return False
    for row in sheet.iter_rows(min_row=2):
        cells = [c.value for c in row]
        entity = _cell_norm(cells, idx["entity"])
        if entity not in ("keyword", "product targeting"):
            continue
        if want_kw and _cell_norm(cells, idx["keyword_text"]) != want_kw:
            continue
        if want_campaign and idx["campaign_name"] is not None and _cell_norm(cells, idx["campaign_name"]) != want_campaign:
            continue
        if want_ad_group and idx["ad_group_name"] is not None and _cell_norm(cells, idx["ad_group_name"]) != want_ad_group:
            continue
        row[idx["bid"]].value = _dollars(new_bid_cents)
        row[idx["operation"]].value = "Update"  # Amazon operation names are capitalized
        return True
    return False


def _find_sibling(sheet, idx, want_campaign, want_ad_group) -> Optional[list]:
    """Return cell-values of any row in the same ad group, to copy IDs from."""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if want_campaign and idx["campaign_name"] is not None and _cell_norm(row, idx["campaign_name"]) != want_campaign:
            continue
        if want_ad_group and idx["ad_group_name"] is not None and _cell_norm(row, idx["ad_group_name"]) != want_ad_group:
            continue
        return list(row)
    return None


def _append_row(sheet, header, idx, rec: Recommendation) -> bool:
    br = rec.bulk_row
    want_campaign = _norm_key(br.get("campaign_name"))
    want_ad_group = _norm_key(br.get("ad_group_name"))
    sibling = _find_sibling(sheet, idx, want_campaign, want_ad_group)
    if sibling is None:
        # Without the campaign/ad-group IDs we can't create a valid row.
        return False

    new_row = [""] * len(header)
    # Copy structural identity from the sibling (IDs, names, product).
    for key in ("campaign_id", "ad_group_id", "campaign_name", "ad_group_name", "product"):
        col = idx[key]
        if col is not None and col < len(sibling):
            new_row[col] = sibling[col]

    is_negative = br["action"] == "create_negative"
    if idx["entity"] is not None:
        new_row[idx["entity"]] = "Negative Keyword" if is_negative else "Keyword"
    if idx["operation"] is not None:
        new_row[idx["operation"]] = "create"
    if idx["keyword_text"] is not None:
        new_row[idx["keyword_text"]] = br.get("keyword_text", "")
    if idx["match_type"] is not None:
        new_row[idx["match_type"]] = "negativeExact" if is_negative else "exact"
    if not is_negative and idx["bid"] is not None and br.get("new_bid_cents"):
        new_row[idx["bid"]] = _dollars(br["new_bid_cents"])

    sheet.append(new_row)
    return True
