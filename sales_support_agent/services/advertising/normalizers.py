"""Tolerant parsers for the Amazon CSV/XLSX exports the audit ingests.

Same philosophy as cashflow's normalizers: absorb varying header spellings via
alias lists, clean values with typed helpers, and NEVER raise — a malformed row
is skipped, not fatal. Each normalizer returns canonical dataclasses from
schema.py.

Supported inputs:
  * Amazon Ads bulk-operations file (XLSX, multi-sheet) -> SP/SB/SD AdRows
  * Sponsored Products Search Term report (CSV)         -> search_term AdRows
  * Business Report: Detail Page Sales & Traffic (CSV)  -> SalesRows
  * Brand Analytics Search Query Performance (CSV)       -> MarketRows
  * DSP performance export (CSV)                         -> DSP AdRows
  * External costs (CSV; Meta/TikTok/influencer)         -> ExternalCostRows
"""

from __future__ import annotations

import csv
import io
import logging
import re
from typing import Iterable, Optional

from sales_support_agent.services.advertising.schema import (
    AdRow,
    ExternalCostRow,
    MarketRow,
    SalesRow,
    parse_bps,
    parse_cents,
    parse_int,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------


def _norm_key(key: object) -> str:
    return str(key or "").strip().lower().replace("﻿", "")


def _lookup(row: dict) -> dict:
    """Return a case/space-insensitive view of a row keyed by normalized name."""
    return {_norm_key(k): ("" if v is None else v) for k, v in row.items()}


def _get(view: dict, *names: str) -> str:
    """First non-empty value across candidate header spellings (substring-
    tolerant): an alias matches a column if it equals it or is contained in it.
    """
    for name in names:
        target = _norm_key(name)
        if target in view and str(view[target]).strip():
            return str(view[target]).strip()
    # substring fallback — handles "spend(usd)" vs "spend", date-suffixed cols
    for name in names:
        target = _norm_key(name)
        for key, value in view.items():
            if target and target in key and str(value).strip():
                return str(value).strip()
    return ""


def _decode(file_bytes: bytes) -> str:
    return file_bytes.decode("utf-8-sig", errors="replace")


def _looks_like_xlsx(file_bytes: bytes) -> bool:
    return file_bytes[:2] == b"PK"  # XLSX is a ZIP container (PK\x03\x04)


def _unwrap_id(value: str) -> str:
    """Amazon report exports wrap IDs as Excel formula text: ="123456". Strip it."""
    v = (value or "").strip()
    if v.startswith('="') and v.endswith('"'):
        return v[2:-1]
    return v.strip('"=')


def _read_csv_rows(file_bytes: bytes, header_hint: Iterable[str]) -> list[dict]:
    """Parse CSV bytes into a list of dict rows, tolerating a preamble before
    the real header. The header row is the first row containing any hint token.
    """
    text = _decode(file_bytes)
    if not text.strip():
        return []
    try:
        raw_lines = list(csv.reader(io.StringIO(text)))
    except (csv.Error, ValueError):
        # Not real CSV (e.g. a binary .xlsx misrouted here) — never raise.
        return []
    if not raw_lines:
        return []
    hints = {_norm_key(h) for h in header_hint}
    header_idx = 0
    for idx, line in enumerate(raw_lines[:15]):
        cells = {_norm_key(c) for c in line}
        if any(any(h in c for c in cells) for h in hints):
            header_idx = idx
            break
    header = [str(c).strip() for c in raw_lines[header_idx]]
    rows: list[dict] = []
    for line in raw_lines[header_idx + 1:]:
        if not any(str(c).strip() for c in line):
            continue
        # pad/truncate to header width
        padded = list(line) + [""] * (len(header) - len(line))
        rows.append(dict(zip(header, padded[: len(header)])))
    return rows


# ---------------------------------------------------------------------------
# Ads bulk-operations file (XLSX) — SP / SB / SD
# ---------------------------------------------------------------------------

# Amazon "Entity" column value -> our entity_level
_ENTITY_MAP = {
    "campaign": "campaign",
    "ad group": "ad_group",
    "keyword": "keyword",
    "product targeting": "target",
    "product ad": "product_ad",
    "product_ad": "product_ad",
    "negative keyword": "negative_keyword",
    "campaign negative keyword": "negative_keyword",
    "negative product targeting": "negative_target",
}


def _ad_type_from_sheet(sheet_name: str) -> Optional[str]:
    name = sheet_name.lower()
    if "sponsored product" in name or name.startswith("sp "):
        return "SP"
    # "Sponsored Brands Campaigns" AND "SB Multi Ad Group Campaigns" (the latter
    # holds 400+ SB keywords and was being skipped by the old check).
    if "sponsored brand" in name or name.startswith("sb "):
        return "SB"
    if "sponsored display" in name or name.startswith("sd "):
        return "SD"
    return None


def normalize_bulk_xlsx(file_bytes: bytes) -> list[AdRow]:
    """Parse an Amazon Ads bulk-operations workbook into AdRows across SP/SB/SD.

    Only performance-bearing entity rows (campaign/ad group/keyword/target/
    product ad) are returned; structural rows without metrics are skipped. The
    full original row is retained in AdRow.raw for traceability.
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - dependency guaranteed in prod
        logger.error("[advertising] openpyxl not installed; cannot parse bulk XLSX")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        logger.exception("[advertising] failed to open bulk XLSX")
        return []

    out: list[AdRow] = []
    for sheet in wb.worksheets:
        ad_type = _ad_type_from_sheet(sheet.title)
        if ad_type is None:
            continue
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        header = [str(h).strip() if h is not None else "" for h in header]
        for values in rows_iter:
            row = dict(zip(header, values))
            view = _lookup(row)
            entity_raw = _norm_key(_get(view, "Entity"))
            level = _ENTITY_MAP.get(entity_raw)
            if level in (None, "negative_keyword", "negative_target"):
                continue
            impressions = parse_int(_get(view, "Impressions"))
            clicks = parse_int(_get(view, "Clicks"))
            spend = parse_cents(_get(view, "Spend"))
            if impressions == 0 and clicks == 0 and spend == 0:
                # No performance signal at this row — skip for snapshots.
                continue
            bid_raw = _get(view, "Bid")
            out.append(
                AdRow(
                    ad_type=ad_type,
                    entity_level=level,
                    campaign_name=_get(view, "Campaign Name (Informational only)", "Campaign Name", "Campaign"),
                    ad_group_name=_get(view, "Ad Group Name (Informational only)", "Ad Group Name", "Ad Group"),
                    entity_text=_get(view, "Keyword Text", "Product Targeting Expression", "Resolved Product Targeting Expression", "ASIN", "SKU"),
                    match_type=_get(view, "Match Type"),
                    impressions=impressions,
                    clicks=clicks,
                    spend_cents=spend,
                    sales_cents=parse_cents(_get(view, "Sales", "14 Day Total Sales", "7 Day Total Sales")),
                    orders=parse_int(_get(view, "Orders", "14 Day Total Orders (#)", "7 Day Total Orders (#)")),
                    units=parse_int(_get(view, "Units", "14 Day Total Units (#)", "7 Day Total Units (#)")),
                    bid_cents=parse_cents(bid_raw) if bid_raw else None,
                    raw={str(k): ("" if v is None else str(v)) for k, v in row.items()},
                )
            )
    return out


# ---------------------------------------------------------------------------
# Sponsored Products Search Term report (CSV)
# ---------------------------------------------------------------------------


def report_date_range(file_bytes: bytes) -> str:
    """The 'Date range' value from a new-console ads report (e.g.
    'May 07, 2026 - May 28, 2026'), used to warn when uploaded reports cover
    different windows. Returns '' if absent. Never raises."""
    for row in _read_csv_rows(file_bytes, header_hint=["Date range", "Campaign", "Total cost"]):
        v = _get(_lookup(row), "Date range")
        if v:
            return v
    return ""


def normalize_ads_report_csv(file_bytes: bytes, ad_type: str = "SP") -> list[AdRow]:
    """Unified parser for Amazon Ads performance reports — both the NEW reporting
    console (columns: Campaign name / Ad group name / Search term | Advertised
    product SKU | Targeting, Impressions, Clicks, Total cost, Purchases, Sales,
    Units sold) and legacy per-entity exports (Total cost (USD), Sales (USD),
    Default bid (USD), 7/14 Day totals).

    Each row's entity level is detected from which entity column is populated, so
    one function ingests search-term, advertised-product, targeting/keyword, ad-
    group and campaign reports. Rows without any performance signal are skipped.
    """
    rows = (
        _read_xlsx_rows(file_bytes)
        if _looks_like_xlsx(file_bytes)
        else _read_csv_rows(file_bytes, header_hint=["Total cost", "Impressions", "Campaign", "Ad group", "Search term"])
    )
    out: list[AdRow] = []
    for row in rows:
        view = _lookup(row)
        impressions = parse_int(_get(view, "Impressions"))
        clicks = parse_int(_get(view, "Clicks"))
        spend = parse_cents(_get(view, "Total cost (USD)", "Total cost", "Spend", "Cost"))
        if impressions == 0 and clicks == 0 and spend == 0:
            continue

        campaign = _get(view, "Campaign name", "Campaign Name", "Campaign")
        ad_group = _get(view, "Ad group name", "Ad Group Name", "Ad Group")
        search_term = _get(view, "Search term", "Customer Search Term")
        # "Targeting" / "Keyword Text" only — NOT a bare "Keyword(s)" count column.
        targeting = _get(view, "Targeting", "Keyword Text", "Keyword text")
        advertised = _get(view, "Advertised product SKU", "Advertised product ID", "Advertised SKU", "Advertised ASIN")
        adv_asin = _get(view, "Advertised product ID", "Advertised ASIN")
        match_type = _get(view, "Match type", "Match Type")

        if search_term:
            level, text = "search_term", search_term
        elif targeting:
            level, text = ("keyword" if match_type else "target"), targeting
        elif advertised:
            level, text = "product_ad", advertised
        elif ad_group and not campaign:
            level, text = "ad_group", ad_group
        elif ad_group and campaign:
            # campaign + ad group present but no finer entity -> ad-group-level report
            level, text = "ad_group", ad_group
        else:
            level, text = "campaign", campaign or ad_group

        bid_raw = _get(view, "Default bid (USD)", "Default bid", "Bid", "CPC (USD)", "CPC")
        raw = {str(k): ("" if v is None else str(v)) for k, v in row.items()}
        # Canonicalize the advertised ASIN so brand-scoping (brand._ad_asin) and
        # the ASIN scorecard find it regardless of the report's column spelling
        # ("Advertised ASIN" in legacy exports vs "Advertised product ID").
        if adv_asin and not raw.get("Advertised product ID"):
            raw["Advertised product ID"] = adv_asin
        out.append(
            AdRow(
                ad_type=ad_type,
                entity_level=level,
                campaign_name=campaign,
                ad_group_name=ad_group,
                campaign_id=_unwrap_id(_get(view, "Campaign ID")),
                ad_group_id=_unwrap_id(_get(view, "Ad group ID", "Ad Group ID")),
                entity_text=text,
                match_type=match_type,
                impressions=impressions,
                clicks=clicks,
                spend_cents=spend,
                sales_cents=parse_cents(_get(view, "Sales (USD)", "Sales", "14 Day Total Sales", "7 Day Total Sales", "Total Sales")),
                orders=parse_int(_get(view, "Purchases", "Orders", "7 Day Total Orders (#)", "14 Day Total Orders (#)")),
                units=parse_int(_get(view, "Units sold", "Units", "7 Day Total Units (#)", "14 Day Total Units (#)")),
                bid_cents=parse_cents(bid_raw) if bid_raw else None,
                raw=raw,
            )
        )
    return out


def normalize_search_term_csv(file_bytes: bytes, ad_type: str = "SP") -> list[AdRow]:
    """Back-compat alias — the unified report normalizer detects search-term rows."""
    return normalize_ads_report_csv(file_bytes, ad_type=ad_type)


def normalize_bulk_keywords(
    file_bytes: bytes, brand_asins: set, other_asins: Optional[set] = None
) -> list[AdRow]:
    """Stream the SP Campaigns sheet of an Amazon Bulk Operations workbook and
    return keyword rows (with Keyword ID, current Bid, and performance) for the
    brand — scoped via each campaign's advertised ASINs (from its Product Ad
    rows). Keeps ONLY campaigns that advertise a brand ASIN and NO other-brand
    ASIN, so a bid change can never touch another brand.

    Memory-safe: read-only streaming with reset_dimensions() (these big Amazon
    exports ship a bogus <dimension>, which otherwise makes read-only yield 0
    rows). Never raises."""
    other_asins = {a.upper() for a in (other_asins or set())}
    brand_asins = {a.upper() for a in (brand_asins or set())}
    if not brand_asins:
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        logger.exception("[advertising] failed to open bulk workbook for keyword scan")
        return []

    ws = None
    for sheet in wb.worksheets:
        if _ad_type_from_sheet(sheet.title) == "SP":
            ws = sheet
            break
    if ws is None:
        wb.close()
        return []
    ws.reset_dimensions()

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []
    idx = {name: (header.index(name) if name in header else None) for name in (
        "Entity", "Campaign ID", "Ad Group ID", "Keyword ID", "Bid", "Keyword Text",
        "Match Type", "ASIN (Informational only)", "Campaign Name (Informational only)",
        "Ad Group Name (Informational only)", "Impressions", "Clicks", "Spend", "Sales",
        "Orders", "Units")}

    def g(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    campaign_asins: dict[str, set] = {}
    keyword_buf: list[tuple] = []
    for row in rows_iter:
        entity = g(row, "Entity")
        cid = str(g(row, "Campaign ID") or "").strip()
        if entity == "Product Ad":
            asin = str(g(row, "ASIN (Informational only)") or "").strip().upper()
            if cid and asin:
                campaign_asins.setdefault(cid, set()).add(asin)
        elif entity == "Keyword":
            keyword_buf.append(row)
    wb.close()

    out: list[AdRow] = []
    for row in keyword_buf:
        cid = str(g(row, "Campaign ID") or "").strip()
        asins = campaign_asins.get(cid, set())
        # brand-only campaigns only: advertises a brand ASIN, no other-brand ASIN.
        if not (asins & brand_asins) or (asins & other_asins):
            continue
        bid = _get(_lookup({"Bid": g(row, "Bid")}), "Bid")
        out.append(AdRow(
            ad_type="SP", entity_level="keyword",
            campaign_name=str(g(row, "Campaign Name (Informational only)") or ""),
            ad_group_name=str(g(row, "Ad Group Name (Informational only)") or ""),
            campaign_id=cid,
            ad_group_id=str(g(row, "Ad Group ID") or "").strip(),
            keyword_id=str(g(row, "Keyword ID") or "").strip(),
            entity_text=str(g(row, "Keyword Text") or ""),
            match_type=str(g(row, "Match Type") or ""),
            impressions=parse_int(g(row, "Impressions")),
            clicks=parse_int(g(row, "Clicks")),
            spend_cents=parse_cents(g(row, "Spend")),
            sales_cents=parse_cents(g(row, "Sales")),
            orders=parse_int(g(row, "Orders")),
            units=parse_int(g(row, "Units")),
            bid_cents=parse_cents(bid) if bid else None,
        ))
    return out


def normalize_bulk_sb(
    file_bytes: bytes, brand_asins: set, other_asins: Optional[set] = None
) -> list[AdRow]:
    """Stream the Sponsored Brands sheets of a Bulk Operations workbook and return
    SB **keyword** and **product-targeting** rows (Keyword/Targeting ID, current
    Bid, performance) for the brand — the only source of SB entity-level data
    (the SB reports are campaign/search-term level). SB campaigns carry no per-row
    ASIN, so scope is taken from each campaign's creative ASINs (Creative ASINs /
    Landing Page ASINs); a campaign with brand creative ASINs and no other-brand
    ASIN is kept. Campaigns with no resolvable ASINs are kept only when there are
    no other-brand ASINs to protect (a single-brand / full-account run).
    Memory-safe; never raises."""
    other_asins = {a.upper() for a in (other_asins or set())}
    brand_asins = {a.upper() for a in (brand_asins or set())}
    if not brand_asins:
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        logger.exception("[advertising] failed to open bulk workbook for SB scan")
        return []

    out: list[AdRow] = []
    for sheet in wb.worksheets:
        if _ad_type_from_sheet(sheet.title) != "SB":
            continue
        try:
            sheet.reset_dimensions()
        except Exception:  # noqa: BLE001
            pass
        it = sheet.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        idx = {n: (header.index(n) if n in header else None) for n in (
            "Entity", "Campaign ID", "Ad Group ID", "Keyword ID", "Product Targeting ID",
            "Product Targeting Expression", "Bid", "Keyword Text", "Match Type",
            "Creative ASINs", "Landing Page ASINs",
            "Campaign Name (Informational only)", "Ad Group Name (Informational only)",
            "Impressions", "Clicks", "Spend", "Sales", "Orders", "Units")}

        def g(row, name, _idx=idx):
            i = _idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        campaign_asins: dict[str, set] = {}
        entity_buf: list[tuple] = []
        for row in it:
            entity = str(g(row, "Entity") or "")
            cid = str(g(row, "Campaign ID") or "").strip()
            asin_text = f"{g(row, 'Creative ASINs') or ''} {g(row, 'Landing Page ASINs') or ''}".upper()
            found = set(_ASIN_RE.findall(asin_text))
            if cid and found:
                campaign_asins.setdefault(cid, set()).update(found)
            if entity in ("Keyword", "Product Targeting"):
                entity_buf.append(row)

        for row in entity_buf:
            cid = str(g(row, "Campaign ID") or "").strip()
            asins = campaign_asins.get(cid, set())
            if asins & other_asins:
                continue  # touches another brand — never edit
            if not (asins & brand_asins) and (asins or other_asins):
                continue  # only other/unknown ASINs in a multi-brand run
            entity = str(g(row, "Entity") or "")
            bid = _get(_lookup({"Bid": g(row, "Bid")}), "Bid")
            is_kw = entity == "Keyword"
            out.append(AdRow(
                ad_type="SB",
                entity_level="keyword" if is_kw else "target",
                bulk_sheet=sheet.title,
                campaign_name=str(g(row, "Campaign Name (Informational only)") or ""),
                ad_group_name=str(g(row, "Ad Group Name (Informational only)") or ""),
                campaign_id=cid,
                ad_group_id=str(g(row, "Ad Group ID") or "").strip(),
                keyword_id=str(g(row, "Keyword ID") or "").strip() if is_kw else "",
                target_id=str(g(row, "Product Targeting ID") or "").strip() if not is_kw else "",
                entity_text=str((g(row, "Keyword Text") if is_kw else g(row, "Product Targeting Expression")) or ""),
                match_type=str(g(row, "Match Type") or ""),
                impressions=parse_int(g(row, "Impressions")),
                clicks=parse_int(g(row, "Clicks")),
                spend_cents=parse_cents(g(row, "Spend")),
                sales_cents=parse_cents(g(row, "Sales")),
                orders=parse_int(g(row, "Orders")),
                units=parse_int(g(row, "Units")),
                bid_cents=parse_cents(bid) if bid else None,
            ))
    wb.close()
    return out


def _norm_target_expr(s: object) -> str:
    """Canonicalize a product-targeting expression for matching across sources.
    Amazon's performance reports write ASIN targets as asin-expanded="B0..."
    while the bulk file writes asin="B0..."; fold them together. Auto expressions
    (close-match / loose-match / substitutes / complements) need only lowercasing."""
    v = str(s or "").strip().lower()
    return v.replace('asin-expanded=', 'asin=')


def bulk_name_id_map(file_bytes: bytes) -> dict:
    """Stream an Amazon Bulk Operations workbook and return name→ID maps:

        {"campaign":  {campaign_name: campaign_id},
         "ad_group":  {(campaign_name, ad_group_name): ad_group_id},
         "keyword":   {(campaign_name, ad_group_name, keyword_text, match_type): keyword_id},
         "target":    {(campaign_name, ad_group_name, expression): product_targeting_id}}

    Built from every SP/SB/SD entity row (Campaign, Ad Group, Keyword, Product
    Targeting), so it covers auto/broad ad groups that hold no keyword rows and
    auto-target expressions (close-match/loose-match/substitutes/complements) +
    ASIN targets. Names/expressions are lowercased/stripped for matching. Used to
    backfill IDs onto ID-less performance reports so their actions become
    apply-ready. Memory-safe (the reset_dimensions() gotcha); never raises."""
    out: dict = {"campaign": {}, "ad_group": {}, "keyword": {}, "target": {}}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        logger.exception("[advertising] failed to open bulk workbook for name→ID map")
        return out

    def nk(v: object) -> str:
        return str(v or "").strip().lower()

    def cid(v: object) -> str:
        return str(v or "").strip()

    for sheet in wb.worksheets:
        if _ad_type_from_sheet(sheet.title) is None:
            continue
        try:
            sheet.reset_dimensions()
        except Exception:  # noqa: BLE001
            pass
        it = sheet.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        idx = {n: (header.index(n) if n in header else None) for n in (
            "Campaign ID", "Ad Group ID", "Keyword ID", "Keyword Text", "Match Type",
            "Product Targeting ID", "Product Targeting Expression",
            "Campaign Name (Informational only)", "Ad Group Name (Informational only)")}

        def g(row, name, _idx=idx):
            i = _idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        for row in it:
            cn, an = nk(g(row, "Campaign Name (Informational only)")), nk(g(row, "Ad Group Name (Informational only)"))
            c_id, a_id, k_id = cid(g(row, "Campaign ID")), cid(g(row, "Ad Group ID")), cid(g(row, "Keyword ID"))
            kt, mt = nk(g(row, "Keyword Text")), nk(g(row, "Match Type"))
            t_id, expr = cid(g(row, "Product Targeting ID")), _norm_target_expr(g(row, "Product Targeting Expression"))
            if cn and c_id:
                out["campaign"].setdefault(cn, c_id)
            if cn and an and a_id:
                out["ad_group"].setdefault((cn, an), a_id)
            if cn and an and kt and k_id:
                out["keyword"].setdefault((cn, an, kt, mt), k_id)
            if cn and an and expr and t_id:
                out["target"].setdefault((cn, an, expr), t_id)
    wb.close()
    return out


def backfill_entity_ids(rows: "list[AdRow]", idmap: dict) -> int:
    """Legacy Amazon "data export" Ads reports carry NO Campaign/Ad Group/Keyword
    IDs, so their harvested keywords, negatives and bid changes can't be written
    to Amazon's apply sheet. Using a bulk_name_id_map(), backfill those IDs onto
    the report rows by matching campaign + ad-group name (and keyword text +
    match type for keyword IDs). Mutates rows in place; returns rows enriched."""
    if not idmap:
        return 0
    camp, adg, kw, tgt = (idmap.get("campaign", {}), idmap.get("ad_group", {}),
                          idmap.get("keyword", {}), idmap.get("target", {}))

    def nk(s: str) -> str:
        return (s or "").strip().lower()

    enriched = 0
    for r in rows:
        cn, an = nk(r.campaign_name), nk(r.ad_group_name)
        changed = False
        if not r.campaign_id and cn in camp:
            r.campaign_id = camp[cn]; changed = True
        if not r.ad_group_id and (cn, an) in adg:
            r.ad_group_id = adg[(cn, an)]; changed = True
        # Keyword ID (managed keywords) — for keyword/target rows that are real keywords.
        if not r.keyword_id and r.entity_level in ("keyword", "target"):
            k = (cn, an, nk(r.entity_text), nk(r.match_type))
            if k in kw:
                r.keyword_id = kw[k]; changed = True
        # Product Targeting ID — auto-target expressions (close/loose-match,
        # substitutes, complements) + ASIN targets. The report writes ASIN targets
        # as asin-expanded="…"; the bulk file as asin="…" — canonicalize both.
        if not r.keyword_id and not r.target_id and tgt:
            t = (cn, an, _norm_target_expr(r.entity_text))
            if t in tgt:
                r.target_id = tgt[t]; changed = True
        enriched += 1 if changed else 0
    return enriched


_ASIN_RE = re.compile(r"B0[A-Z0-9]{8}")


def bulk_sp_home_by_asin(file_bytes: bytes) -> dict:
    """For each advertised ASIN, the best Sponsored Products 'home' to add a
    harvested exact keyword to: the SP campaign that advertises the ASIN, and
    within it the ad group holding the most existing keywords (i.e. the managed/
    exact ad group, not an auto one). Returns
        {asin: (campaign_id, ad_group_id, campaign_name, ad_group_name)}.
    Memory-safe; never raises."""
    from collections import Counter, defaultdict
    out: dict = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        logger.exception("[advertising] failed to open bulk workbook for SP-home map")
        return out

    asin_camps: dict = defaultdict(set)
    kwcount: Counter = Counter()
    names: dict = {}
    for sheet in wb.worksheets:
        if _ad_type_from_sheet(sheet.title) != "SP":
            continue
        try:
            sheet.reset_dimensions()
        except Exception:  # noqa: BLE001
            pass
        it = sheet.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        idx = {h: (header.index(h) if h in header else None) for h in (
            "Entity", "Campaign ID", "Ad Group ID", "ASIN (Informational only)",
            "Campaign Name (Informational only)", "Ad Group Name (Informational only)")}

        def g(row, name, _idx=idx):
            i = _idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        for row in it:
            entity = str(g(row, "Entity") or "")
            cid = str(g(row, "Campaign ID") or "").strip()
            aid = str(g(row, "Ad Group ID") or "").strip()
            if entity == "Product Ad":
                asin = str(g(row, "ASIN (Informational only)") or "").strip().upper()
                if asin and cid:
                    asin_camps[asin].add(cid)
            elif entity == "Keyword" and cid and aid:
                kwcount[(cid, aid)] += 1
                names[(cid, aid)] = (str(g(row, "Campaign Name (Informational only)") or ""),
                                     str(g(row, "Ad Group Name (Informational only)") or ""))
    wb.close()

    for asin, camps in asin_camps.items():
        cands = [(cid, aid) for (cid, aid) in kwcount if cid in camps]
        if not cands:
            continue
        cid, aid = max(cands, key=lambda k: kwcount[k])
        cn, an = names.get((cid, aid), ("", ""))
        out[asin] = (cid, aid, cn, an)
    return out


def redirect_harvests_to_sp(recs: list, sp_home: dict) -> int:
    """Account-manager preference: a winning search term discovered in a
    Sponsored Brands (or otherwise unresolved) campaign should be harvested into
    the **Sponsored Products** campaign that advertises the same ASIN. For each
    create_keyword rec still missing campaign/ad-group IDs, pull the ASIN from the
    source campaign name and repoint it at that ASIN's SP home (bulk_sp_home_by_asin).
    Left untouched (stays in the burn list, not the apply sheet) when no SP home
    exists. Mutates rec.bulk_row in place; returns the count redirected."""
    if not sp_home:
        return 0
    n = 0
    for r in recs:
        br = getattr(r, "bulk_row", None) or {}
        if br.get("action") != "create_keyword" or (br.get("campaign_id") and br.get("ad_group_id")):
            continue
        m = _ASIN_RE.search((br.get("campaign_name") or "").upper())
        if not m:
            continue
        home = sp_home.get(m.group(0))
        if not home:
            continue
        cid, aid, cn, an = home
        br["campaign_id"], br["ad_group_id"] = cid, aid
        br["campaign_name"], br["ad_group_name"] = cn, an
        br["ad_type"] = "SP"
        r.is_bulk_actionable = True
        n += 1
    return n


_AUTO_EXPRS = {"close-match", "loose-match", "substitutes", "complements"}


def enforce_targeting_type(recs: list, file_bytes: bytes) -> int:
    """A Sponsored Products ad group is EITHER keyword-targeted OR product-
    targeted — never both — and an **auto** ad group takes no manual entities at
    all. Auto/discovery ad groups surface both keyword-like and ASIN search
    terms, so harvesting them back into that ad group mixes types (and targets an
    auto ad group), which Amazon rejects.

    For every positive harvest (`create_keyword`), this re-homes it into a MANUAL
    ad group of the matching targeting type for the same ASIN (a keyword ad group
    for keyword harvests, a manual product-targeting ad group for ASIN harvests),
    using the bulk file. If the current ad group is already the right manual type
    it's left alone; if no suitable home exists the harvest is dropped from the
    apply sheet (stays in the burn list). Returns rows re-homed or dropped."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        logger.exception("[advertising] failed to open bulk workbook for targeting-type check")
        return 0
    from collections import Counter, defaultdict

    kw_ag: Counter = Counter()          # (cid,aid) -> # manual keywords
    pt_ag: Counter = Counter()          # (cid,aid) -> # manual product targets
    auto_ag: set = set()                # (cid,aid) with auto targeting
    camp_asins: dict = defaultdict(set)
    for sheet in wb.worksheets:
        if _ad_type_from_sheet(sheet.title) != "SP":
            continue
        try:
            sheet.reset_dimensions()
        except Exception:  # noqa: BLE001
            pass
        it = sheet.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        idx = {h: (header.index(h) if h in header else None) for h in (
            "Entity", "Campaign ID", "Ad Group ID", "ASIN (Informational only)", "Product Targeting Expression")}

        def g(row, name, _idx=idx):
            i = _idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        for row in it:
            e = str(g(row, "Entity") or "")
            cid = str(g(row, "Campaign ID") or "").strip()
            aid = str(g(row, "Ad Group ID") or "").strip()
            if e == "Product Ad":
                a = str(g(row, "ASIN (Informational only)") or "").strip().upper()
                if a and cid:
                    camp_asins[cid].add(a)
            elif e == "Keyword" and cid and aid:
                kw_ag[(cid, aid)] += 1
            elif e == "Product Targeting" and cid and aid:
                expr = str(g(row, "Product Targeting Expression") or "").strip().lower()
                if expr in _AUTO_EXPRS:
                    auto_ag.add((cid, aid))
                else:
                    pt_ag[(cid, aid)] += 1
    wb.close()

    def homes(counter: Counter) -> dict:
        by_asin: dict = defaultdict(list)
        for (cid, aid), n in counter.items():
            for a in camp_asins.get(cid, ()):
                by_asin[a].append((n, cid, aid))
        return {a: max(v)[1:] for a, v in by_asin.items()}

    kw_home, pt_home = homes(kw_ag), homes(pt_ag)
    kw_set, pt_set = set(kw_ag), set(pt_ag)

    def _is_target(text: str) -> bool:
        t = (text or "").strip().lower()
        return bool(re.fullmatch(r"b0[a-z0-9]{8}", t)) or t.startswith(("asin=", "asin-expanded=", "category=", "brand="))

    changed = 0
    for r in recs:
        br = getattr(r, "bulk_row", None) or {}
        if br.get("action") != "create_keyword":
            continue  # only positive harvests define an ad group's targeting type
        is_target = _is_target(br.get("keyword_text", ""))
        cur = (str(br.get("campaign_id") or ""), str(br.get("ad_group_id") or ""))
        # already in a manual ad group of the right type (and not auto)?
        if cur not in auto_ag and ((is_target and cur in pt_set) or (not is_target and cur in kw_set)):
            continue
        asin_m = _ASIN_RE.search((br.get("campaign_name") or "").upper())
        asin = asin_m.group(0) if asin_m else next(iter(camp_asins.get(cur[0], set())), None)
        home = (pt_home if is_target else kw_home).get(asin) if asin else None
        if home:
            br["campaign_id"], br["ad_group_id"] = home
            changed += 1
        else:
            r.is_bulk_actionable = False  # no valid manual home → keep in burn list, not the sheet
            changed += 1
    return changed


def merge_duplicate_entities(ad_rows: "list[AdRow]") -> "tuple[list[AdRow], int]":
    """The SAME keyword/target arrives as multiple rows — once from the
    performance reports and once from the bulk file — often with different click
    counts (different/overlapping windows). Optimizing each row independently
    produces conflicting bid recs for one keyword, and the apply-sheet dedup can
    then keep the WRONG one (a bid computed from a partial slice).

    Collapse to ONE row per resolved entity ID, keeping the **richest-data view**
    (most clicks; ties → most spend) so each keyword is judged exactly once on its
    fullest performance. Rows without an ID (search terms, un-resolved targets)
    pass through untouched. Returns (merged_rows, count_collapsed)."""
    best: dict = {}
    passthrough: list = []
    collapsed = 0
    for r in ad_rows:
        eid = r.keyword_id or r.target_id
        if r.entity_level not in ("keyword", "target") or not eid:
            passthrough.append(r)
            continue
        key = (r.ad_type, r.entity_level, eid)
        cur = best.get(key)
        if cur is None:
            best[key] = r
        else:
            collapsed += 1
            # richest data wins; carry a real bid forward if the winner lacks one.
            winner = max((cur, r), key=lambda x: (x.clicks, x.spend_cents))
            loser = r if winner is cur else cur
            if not winner.bid_cents and loser.bid_cents:
                winner.bid_cents = loser.bid_cents
            best[key] = winner
    return passthrough + list(best.values()), collapsed


def drop_brand_term_harvests(recs: list, brand_name: str) -> int:
    """Don't harvest a search term that contains the brand name — you already bid
    on your own brand, so creating it as a 'new' keyword is redundant AND almost
    always rejected as 'already exists'. Leaves NEGATIVES alone (a competitor
    using your brand is a valid negative). Only drops positive create_keyword
    harvests. Returns the count dropped."""
    brand = _norm_kw_text(brand_name)
    if not brand:
        return 0
    keys = {brand, brand.replace(" ", "")}  # "number 4" and "number4"
    keys = {k for k in keys if len(k) >= 2}
    n = 0
    for r in recs:
        br = getattr(r, "bulk_row", None) or {}
        if br.get("action") != "create_keyword" or not getattr(r, "is_bulk_actionable", False):
            continue
        t = _norm_kw_text(br.get("keyword_text", ""))
        tc = t.replace(" ", "")
        if any(k in t or k in tc for k in keys):
            r.is_bulk_actionable = False
            n += 1
    return n


def _norm_kw_text(s: object) -> str:
    """Normalize keyword text the way Amazon does for duplicate detection: drop
    punctuation (so 'no. 4 shampoo' == 'no 4 shampoo'), lowercase, collapse
    whitespace. Prevents 'already exists!' rejections on punctuation variants."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", str(s or "").lower())).strip()


def _final_target_expr(text: str) -> str:
    """The product-targeting expression a harvest will be written as: a bare ASIN
    becomes asin="B0…"; asin-expanded= becomes asin=. (Matches build_apply_sheet.)"""
    t = (text or "").strip()
    if re.fullmatch(r"(?i)b0[a-z0-9]{8}", t):
        return f'asin="{t}"'
    return re.sub(r"(?i)asin-expanded\s*=", "asin=", t)


def drop_existing_creates(recs: list, file_bytes: bytes) -> int:
    """Amazon rejects — with an Input Error that fails the WHOLE file — any Create
    for a keyword / negative keyword / product target that ALREADY EXISTS in the
    ad group ("…already exists!"). The uploaded bulk file lists every existing
    entity, so drop those creates here. Uses each rec's FINAL ad group, so call
    AFTER enforce_targeting_type. Marks duplicates not-actionable (they remain in
    the burn list). Returns the count dropped."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        logger.exception("[advertising] failed to open bulk workbook for existing-entity scan")
        return 0

    def mt(s: object) -> str:
        return re.sub(r"[^a-z]", "", str(s or "").lower())  # exact / negativeexact / …

    existing_kw: set = set()
    existing_neg: set = set()
    existing_camp_neg: set = set()
    existing_pt: set = set()
    existing_pt_camp: set = set()   # (campaign, expr) — an ASIN target can't repeat in a campaign
    existing_npt: set = set()
    for sheet in wb.worksheets:
        if _ad_type_from_sheet(sheet.title) is None:
            continue
        try:
            sheet.reset_dimensions()
        except Exception:  # noqa: BLE001
            pass
        it = sheet.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        idx = {h: (header.index(h) if h in header else None) for h in (
            "Entity", "Campaign ID", "Ad Group ID", "Keyword Text", "Match Type", "Product Targeting Expression")}

        def g(row, name, _idx=idx):
            i = _idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        for row in it:
            e = str(g(row, "Entity") or "")
            cid = str(g(row, "Campaign ID") or "").strip()
            aid = str(g(row, "Ad Group ID") or "").strip()
            kt = _norm_kw_text(g(row, "Keyword Text"))
            m = mt(g(row, "Match Type"))
            ex = _norm_target_expr(g(row, "Product Targeting Expression"))
            if e == "Keyword" and kt:
                existing_kw.add((cid, aid, kt, m))
            elif e == "Negative Keyword" and kt:
                existing_neg.add((cid, aid, kt, m))
            elif e == "Campaign Negative Keyword" and kt:
                existing_camp_neg.add((cid, kt, m))
            elif e in ("Product Targeting", "Negative Product Targeting") and ex:
                # campaign-negative PT also bars adding the ASIN as a positive target.
                if e == "Negative Product Targeting":
                    existing_npt.add((cid, aid, ex))
                else:
                    existing_pt.add((cid, aid, ex))
                existing_pt_camp.add((cid, ex))
            elif e == "Campaign Negative Product Targeting" and ex:
                existing_pt_camp.add((cid, ex))
    wb.close()

    def _is_target(text: str) -> bool:
        t = (text or "").strip().lower()
        return bool(re.fullmatch(r"b0[a-z0-9]{8}", t)) or t.startswith(("asin=", "asin-expanded=", "category=", "brand="))

    dropped = 0
    for r in recs:
        br = getattr(r, "bulk_row", None) or {}
        action = br.get("action")
        if action not in ("create_keyword", "create_negative") or not getattr(r, "is_bulk_actionable", False):
            continue
        cid = str(br.get("campaign_id") or "")
        aid = str(br.get("ad_group_id") or "")
        text = br.get("keyword_text", "")
        is_neg = action == "create_negative"
        if _is_target(text):
            ex = _norm_target_expr(_final_target_expr(text))
            if is_neg:
                exists = (cid, aid, ex) in existing_npt
            else:
                # already targeted in this ad group, or anywhere in the campaign
                # (incl. as a campaign-negative) — any of these rejects the add.
                exists = (cid, aid, ex) in existing_pt or (cid, ex) in existing_pt_camp
        else:
            t = _norm_kw_text(text)
            if is_neg:
                exists = (cid, aid, t, "negativeexact") in existing_neg or (cid, t, "negativeexact") in existing_camp_neg
            else:
                exists = (cid, aid, t, "exact") in existing_kw
        if exists:
            r.is_bulk_actionable = False
            dropped += 1
    return dropped


# ---------------------------------------------------------------------------
# Business Report: Detail Page Sales & Traffic (CSV)
# ---------------------------------------------------------------------------


def normalize_business_report_csv(file_bytes: bytes) -> list[SalesRow]:
    rows = _read_csv_rows(file_bytes, header_hint=["Sessions", "ASIN", "Units Ordered"])
    out: list[SalesRow] = []
    for row in rows:
        view = _lookup(row)
        asin = _get(view, "(Child) ASIN", "Child ASIN", "ASIN", "(Parent) ASIN")
        sku = _get(view, "SKU")
        if not asin and not sku:
            continue
        out.append(
            SalesRow(
                asin=asin,
                sku=sku,
                title=_get(view, "Title", "Product Name"),
                sessions=parse_int(_get(view, "Sessions - Total", "Sessions – Total", "Sessions")),
                page_views=parse_int(_get(view, "Page Views - Total", "Page Views – Total", "Page Views")),
                units=parse_int(_get(view, "Units Ordered")),
                ordered_product_sales_cents=parse_cents(_get(view, "Ordered Product Sales")),
                buy_box_pct_bps=parse_bps(_get(view, "Featured Offer (Buy Box) Percentage", "Buy Box Percentage")),
                conversion_bps=parse_bps(_get(view, "Unit Session Percentage")),
                raw={str(k): str(v) for k, v in row.items()},
            )
        )
    return out


# ---------------------------------------------------------------------------
# Brand Analytics Search Query Performance (CSV)
# ---------------------------------------------------------------------------


def normalize_sqp_csv(file_bytes: bytes) -> list[MarketRow]:
    rows = _read_csv_rows(file_bytes, header_hint=["Search Query", "Search Query Volume"])
    out: list[MarketRow] = []
    for row in rows:
        view = _lookup(row)
        query = _get(view, "Search Query")
        if not query:
            continue
        out.append(
            MarketRow(
                search_query=query,
                asin=_get(view, "ASIN"),
                search_query_volume=parse_int(_get(view, "Search Query Volume")),
                impressions_total=parse_int(_get(view, "Impressions: Total Count", "Impressions Total Count")),
                impression_share_bps=parse_bps(_get(view, "Impressions: ASIN Share %", "Impressions ASIN Share")),
                clicks_total=parse_int(_get(view, "Clicks: Total Count", "Clicks Total Count")),
                click_share_bps=parse_bps(_get(view, "Clicks: ASIN Share %", "Clicks ASIN Share")),
                purchases_total=parse_int(_get(view, "Purchases: Total Count", "Purchases Total Count")),
                purchase_share_bps=parse_bps(_get(view, "Purchases: ASIN Share %", "Purchases ASIN Share")),
                raw={str(k): str(v) for k, v in row.items()},
            )
        )
    return out


# ---------------------------------------------------------------------------
# DSP performance (CSV) — campaign-level
# ---------------------------------------------------------------------------


def normalize_dsp_csv(file_bytes: bytes) -> list[AdRow]:
    rows = _read_csv_rows(file_bytes, header_hint=["Campaign", "Total Cost", "Impressions"])
    out: list[AdRow] = []
    for row in rows:
        view = _lookup(row)
        campaign = _get(view, "Campaign Name", "Campaign", "Order")
        if not campaign:
            continue
        out.append(
            AdRow(
                ad_type="DSP",
                entity_level="campaign",
                campaign_name=campaign,
                entity_text=campaign,
                impressions=parse_int(_get(view, "Impressions")),
                clicks=parse_int(_get(view, "Clicks", "Click-throughs")),
                spend_cents=parse_cents(_get(view, "Total Cost", "Spend", "Cost")),
                sales_cents=parse_cents(_get(view, "Total Sales", "14 Day Total Sales", "Sales")),
                orders=parse_int(_get(view, "Total Orders", "Orders", "Purchases")),
                units=parse_int(_get(view, "Total Units", "Units")),
                raw={str(k): str(v) for k, v in row.items()},
            )
        )
    return out


# ---------------------------------------------------------------------------
# External costs (CSV) — Meta / TikTok / influencer commissions
# ---------------------------------------------------------------------------

_CHANNEL_ALIASES = {
    "meta": "meta", "facebook": "meta", "fb": "meta", "instagram": "meta",
    "tiktok": "tiktok", "tik tok": "tiktok",
    "influencer": "influencer", "creator": "influencer", "affiliate": "influencer",
    "google": "google", "adwords": "google",
}


# Header tokens that represent a per-unit cost component (summed for landed cost).
_COST_TERMS = (
    "cogs", "cost of goods", "unit cost", "landed cost", "fba fee", "fulfillment fee",
    "referral fee", "amz fee", "amazon fee", "freight",
)


def _read_xlsx_rows(file_bytes: bytes) -> list[dict]:
    """First sheet of a workbook as a list of header->value dicts. Never raises."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        try:
            # Amazon exports ship a bogus <dimension> (e.g. A1:A1) that makes
            # read_only iteration yield zero rows; reset it to the real extent.
            ws.reset_dimensions()
        except Exception:  # noqa: BLE001
            pass
        header: Optional[list[str]] = None
        out: list[dict] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if header is None:
                if any(c.strip() for c in cells):
                    header = [c.strip() for c in cells]
                continue
            if not any(c.strip() for c in cells):
                continue
            out.append(dict(zip(header, cells)))
        return out
    except Exception:
        logger.exception("[advertising] failed reading COGS xlsx")
        return []


def _unit_cost_cents(view: dict) -> int:
    """Landed cost = one value per distinct cost component. Margin sheets repeat
    fee columns per price-point (AMZ Fee / AMZ Fee2 / AMZ Fee3); we take the first
    match for each term so they aren't summed multiple times. Fees are stored
    negative, so absolute values are used."""
    found: dict[str, int] = {}
    for key, value in view.items():
        for term in _COST_TERMS:
            if term in key and term not in found:
                found[term] = abs(parse_cents(value))
                break
    total = sum(found.values())
    if total == 0:
        total = abs(parse_cents(_get(view, "Cost", "Price Paid")))
    return total


def normalize_cogs_csv(file_bytes: bytes, sales_rows: "list[SalesRow] | None" = None) -> dict:
    """Parse a per-unit cost sheet (CSV or XLSX) into
    {"asin": {asin: cents}, "sku": {sku: cents}}.

    Keyed by ASIN or SKU when present. For margin sheets keyed by product name
    (no ASIN), we best-effort map product names to ASINs using the Business
    Report titles in `sales_rows`. Cost is the sum of all cost components (COGS +
    FBA/referral/AMZ fees + freight). Never raises."""
    if _looks_like_xlsx(file_bytes):
        rows = _read_xlsx_rows(file_bytes)
    else:
        rows = _read_csv_rows(file_bytes, header_hint=["ASIN", "SKU", "COGS", "Cost", "Unit Cost", "Margin"])

    by_asin: dict[str, int] = {}
    by_sku: dict[str, int] = {}
    source: dict[str, str] = {}  # asin -> where its cost came from (for review)
    named: list[tuple[str, int, str]] = []  # (product label, cost, status)
    for row in rows:
        view = _lookup(row)
        asin = _get(view, "ASIN", "(Child) ASIN", "Child ASIN")
        sku = _get(view, "SKU", "Seller SKU", "MSKU")
        cost = _unit_cost_cents(view)
        if cost <= 0:
            continue
        if asin:
            by_asin[asin] = cost
            source[asin] = "ASIN column (exact)"
        elif sku:
            by_sku[sku] = cost
        else:
            label = (_get(view, "Product Family") + " " + _get(view, "Product", "Product Name", "Title")).strip()
            if label:
                named.append((label, cost, _norm_key(_get(view, "Status"))))

    # Name -> ASIN mapping for margin sheets without an ASIN column.
    if named and sales_rows:
        for s in sales_rows:
            if not s.asin or s.asin in by_asin:
                continue
            match = _match_named_cost(s.title, named)
            if match is not None:
                by_asin[s.asin] = match[0]
                source[s.asin] = f"name-matched: {match[1]}"
    return {"asin": by_asin, "sku": by_sku, "source": source}


_SIZE_RE = re.compile(r"(\d+)\s*(?:ct|count|stix|sticks|pack| servings?)", re.IGNORECASE)


def _match_named_cost(title: str, named: list[tuple[str, int, str]]) -> Optional[tuple[int, str]]:
    """Conservatively match a Business Report title to a named cost row: require
    overlap of distinctive word tokens AND, when both carry a size/count, that
    the sizes match. Prefer the 'base price' row. Returns (cents, label) or None."""
    title_l = title.lower()
    title_tokens = {t for t in re.split(r"[^a-z0-9]+", title_l) if len(t) > 2}
    title_size = _SIZE_RE.search(title_l)
    title_size_n = title_size.group(1) if title_size else None

    best: Optional[tuple[int, int, str]] = None  # (score, cost, label)
    for label, cost, status in named:
        label_l = label.lower()
        label_tokens = {t for t in re.split(r"[^a-z0-9]+", label_l) if len(t) > 2}
        overlap = len(title_tokens & label_tokens)
        if overlap < 2:
            continue
        label_size = _SIZE_RE.search(label_l)
        # If both name a size, they must agree; if only one does, allow but rank lower.
        if title_size_n and label_size:
            if label_size.group(1) != title_size_n:
                continue  # sizes conflict -> not the same SKU
            overlap += 2     # size agreement is a strong signal
        score = overlap + (1 if "base" in status else 0)
        if best is None or score > best[0]:
            best = (score, cost, label)
    return (best[1], best[2]) if best else None


def normalize_external_costs_csv(file_bytes: bytes) -> list[ExternalCostRow]:
    rows = _read_csv_rows(file_bytes, header_hint=["Channel", "Amount", "Spend", "Cost"])
    out: list[ExternalCostRow] = []
    for row in rows:
        view = _lookup(row)
        channel_raw = _norm_key(_get(view, "Channel", "Platform", "Source"))
        amount = parse_cents(_get(view, "Amount", "Spend", "Cost", "Commission"))
        if not channel_raw and not amount:
            continue
        channel = _CHANNEL_ALIASES.get(channel_raw, channel_raw or "other")
        cost_type = "commission" if channel == "influencer" else "ad_spend"
        out.append(
            ExternalCostRow(
                channel=channel if channel in {"meta", "tiktok", "influencer", "google"} else "other",
                cost_type=_get(view, "Cost Type", "Type") or cost_type,
                label=_get(view, "Label", "Name", "Campaign") or channel_raw,
                amount_cents=amount,
                note=_get(view, "Note", "Notes"),
            )
        )
    return out
